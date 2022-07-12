# This is a sample Python script.1

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.
import dataclasses

from openpyxl import load_workbook, Workbook
from dataclasses import dataclass

from openpyxl.cell import MergedCell

from MEPItem import MEPItem
from MEPItem2 import MEPItem2


def excel_normalize(name):
    # Use a breakpoint in the code line below to debug your script.
    excel = load_workbook(
        'C:\\Users\\gyujin\\Desktop\\영등포 필립메디컬 증축공사(산출서).xlsx',
        data_only=True)
    # names = excel.get_sheet_names()
    # print(names)
    worksheet = excel['목록별산출서']
    items = []
    for row in worksheet.iter_rows(min_col=4, max_col=13, min_row=6):
        if row[4].value is None:
            continue
        if row[4].value == '명칭':
            continue
        item = MEPItem(
            name=row[4].value,
            standard=row[5].value,
            unit=row[6].value,
            formula=row[1].value,
            unit_formula=row[7].value,
            sum=row[9].value,
        )
        if item.formula is None:
            item.formula = items[-1].formula
        items.append(item)
    print(items.__sizeof__())

    # worksheet2 = excel['산출집계']
    # items2 = []
    # for row in worksheet2.iter_rows(min_col=4, max_col=10, min_row=7):
    #     if row[4].value is None:
    #         continue
    #     if row[6].value == '소 계':
    #         continue
    #     item2 = MEPItem2(
    #         name=row[0].value,
    #         standard=row[1].value,
    #         unit=row[2].value,
    #         sum=row[6].value,
    #     )
    #     items2.append(item2)
    # print(items2)



    # 저장할 엑셀
    new_excel = Workbook()
    new_sheet = new_excel.active
    new_sheet.title = '기계'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    for item in items:
        new_sheet.append(item.to_excel())

    # sheet = new_excel.create_sheet(title='집계표')
    # sheet.append(['중공종', '품명', '규격', '단위', '수량(할증전)'])
    # for item2 in items2:
    #     sheet.append(item2.to_excel())

    # excel = load_workbook(
    #     'C:\\서울시 은평구 역촌동 77-9,47 근생주택 신축공사-통신(산출).xlsm',
    #     data_only=True)
    # # names = excel.get_sheet_names()
    # # print(names)
    # worksheet = excel['목록별산출서']
    # items = []
    # for row in worksheet.iter_rows(min_col=4, max_col=13, min_row=6):
    #     if row[4].value is None:
    #         continue
    #     if row[4].value == '명칭':
    #         continue
    #     item = MEPItem(
    #         name=row[4].value,
    #         standard=row[5].value,
    #         unit=row[6].value,
    #         formula=row[1].value,
    #         unit_formula=row[7].value,
    #         sum=row[9].value,
    #     )
    #     if item.formula is None:
    #         item.formula = items[-1].formula
    #     items.append(item)
    # print(items.__sizeof__())
    #
    # worksheet2 = excel['산출집계']
    # items2 = []
    # for row in worksheet2.iter_rows(min_col=4, max_col=10, min_row=7):
    #     if row[4].value is None:
    #         continue
    #     if row[6].value == '소 계':
    #         continue
    #     item2 = MEPItem2(
    #         name=row[0].value,
    #         standard=row[1].value,
    #         unit=row[2].value,
    #         sum=row[6].value,
    #     )
    #     items2.append(item2)
    # print(items2)
    #
    # for item in items:
    #     new_sheet.append(item.to_excel())
    #
    # for item2 in items2:
    #     sheet.append(item2.to_excel())
    #
    # excel = load_workbook(
    #     'C:\\서울시 은평구 역촌동 77-9,47 근생주택 신축공사-소방(산출).xlsm',
    #     data_only=True)
    # # names = excel.get_sheet_names()
    # # print(names)
    # worksheet = excel['목록별산출서']
    # items = []
    # for row in worksheet.iter_rows(min_col=4, max_col=13, min_row=6):
    #     if row[4].value is None:
    #         continue
    #     if row[4].value == '명칭':
    #         continue
    #     item = MEPItem(
    #         name=row[4].value,
    #         standard=row[5].value,
    #         unit=row[6].value,
    #         formula=row[1].value,
    #         unit_formula=row[7].value,
    #         sum=row[9].value,
    #     )
    #     if item.formula is None:
    #         item.formula = items[-1].formula
    #     items.append(item)
    # print(items.__sizeof__())
    #
    # worksheet2 = excel['산출집계']
    # items2 = []
    # for row in worksheet2.iter_rows(min_col=4, max_col=10, min_row=7):
    #     if row[4].value is None:
    #         continue
    #     if row[6].value == '소 계':
    #         continue
    #     item2 = MEPItem2(
    #         name=row[0].value,
    #         standard=row[1].value,
    #         unit=row[2].value,
    #         sum=row[6].value,
    #     )
    #     items2.append(item2)
    # print(items2)
    #
    # for item in items:
    #     new_sheet.append(item.to_excel())
    #
    # for item2 in items2:
    #     sheet.append(item2.to_excel())

    new_excel.save("C:\\Users\\gyujin\\Desktop\\test.xlsx")


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
