# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.
import dataclasses

from openpyxl import load_workbook, Workbook
from dataclasses import dataclass

from openpyxl.cell import MergedCell

from QuantityItemStandard import QuantityItemStandard
from QuantityItemStandard2 import QuantityItemStandard2
from structure.ItemStandard import ItemStandard


def excel_normalize(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\구조.xlsx',
        data_only=True)
    # names = excel.get_sheet_names()
    # print(names)
    worksheet = excel['부재별산출서']
    items = []
    floor = ""
    ho = ""
    room = ""
    part = ""
    for row in worksheet.iter_rows(min_col=0, max_col=6, min_row=4):
        if ( row[0].value is not None
                and row[1].value is None
                and row[2].value is None
                and row[3].value is None
                and row[4].value is None
                and row[5].value is None
        ):
            ho = row[0].value.split('-')[-1].strip()
            continue

        if ( row[0].value is not None
            and row[1].value is not None):
            floor = row[0].value
            part = row[1].value

        concs = row[3].value.split('-')
        newConc = row[3].value
        if( len(concs) == 3):
            slope = concs[-1].zfill(2)
            newConc = '-'.join([concs[0], concs[1], slope])

        item = ItemStandard(
            floor = floor,
            ho = ho,
            name = row[2].value,
            standard= newConc,
            part = part,
            formula = row[4].value,
            sum = row[5].value,
        )
        print(item.to_excel())
        items.append(item)


    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '전기(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    for item in items:
        new_sheet.append(item.to_excel())


    new_workbook.save("C:\\Users\ckddn\Desktop\구조완성.xlsx")

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/

