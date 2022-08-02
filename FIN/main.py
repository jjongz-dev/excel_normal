import re

from openpyxl import load_workbook, Workbook

from FIN.ItemStandard import ItemStandard

from FIN.ItemStandard2 import ItemStandard2


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\건축.xlsx',
        data_only=True)
    # names = excel.get_sheet_names()
    # print(names)
    worksheet = excel['산출근거집계표']
    items = []
    for row in worksheet.iter_rows(min_col=1, max_col=13, min_row=5):
        # #0값 삭제
        # if ( row[9].value == '0'
        #     and row[12].value == '0'):
        #     continue
        # 산출식 없음 삭제
        if ( row[2].value == '합        계'):
            continue
        # 구조이기 삭제
        if ( row[7].value == '구조이기'):
            continue

        # 0값 삭제
        if (row[12].value == '0' or row[12].value == 0 ):
            continue


        # for cell in row:
        #     print(cell.value, end=', ')

        # print(','.join([
        #         row[0].value,
        #         row[1].value,
        #         row[2].value,
        #         row[3].value,
        #         row[4].value,
        #         row[5].value,
        #         row[6].value,
        #         row[7].value,
        #         row[8].value,
        #         row[9].value,
        #         row[10].value,
        #         row[11].value,]
        #     ),
        #     end=''
        # )
        # for cell in row:
        #     print(cell.col_idx, cell.value)
        item = ItemStandard(
            floor = row[7].value,
            location = '',
            roomname=row[8].value,



            name = row[2].value,
            standard= row[3].value,
            unit=row[4].value,

            type=row[5].value,
            formula = row[9].value,
            sum=row[12].value,
            )
        # print(item.to_excel())
        items.append(item)

    # and item.floor in ['1cmd', '2cmd'] ★ contains 대량

    worksheet2 = excel['동별집계표']
    items2 = []
    constructionWork = ""
    # name = ""
    for row in worksheet2.iter_rows(min_col=1, max_col=5, min_row=4):
        #중공종
        if (row[1].value.__contains__('내  역  삭  제')
        ):
            continue

        if (row[1].value is not None
               and row[4].value is None
        ):
            constructionWork = row[1].value.replace(" ","")


        item2 = ItemStandard2(
            constructionWork = constructionWork,
            name = row[1].value,
            standard = row[2].value,
            unit = row[3].value,
            sum = row[4].value,
            )
        items2.append(item2)

    for item in items:
        #TYPE변경
        type1 = item.type.split('-')
        if (len(type1) == 2):
            item.type = item.type.split('-')[0]

        if (item.type == '토공'):
            item.type = '외부'

        if (item.type == '가설'):
            item.type = '내부'

        #품명변경
        if (item.name.startswith("★")):
            item.name = item.name.replace("★","")

        #창호
        if (item.type == '창호'):
            item.part = item.floor
            item.floor = ''
            item.roomname =''

        #외부 입면 호실정리      일단 하자
        if (item.type == '외부'
                and item.floor.__contains__('정면')):
            item.location = '정면'
            item.roomname = '정면'
            item.floor = ''

        if (item.type == '외부'
            and item.floor.__contains__('배면')

        ):
            item.location = '배면'
            item.roomname = '배면'
            item.floor = ''

        if (item.type == '외부'
            and item.floor.__contains__('좌측면')):
            item.location = '좌측면'
            item.roomname = '좌측면'
            item.floor = ''

        if (item.type == '외부'
            and item.floor.__contains__('우측면')):
            item.location = '우측면'
            item.roomname = '우측면'
            item.floor = ''

        #토공사정리, 기초단열재
        if (item.floor == '토공사' or item.floor == '기초단열재'):
            item.location = '기초하부'
            item.roomname = '기초하부'
            item.floor = 'FT'

        #기본 층정리
        if (item.name in ['가설컨테이너반입', '가설컨테이너반출', '가설수도', '가설전기', '가설울타리설치', '가설울타리해체', '가설출입구설치', '가설출입구해체',
                          '건축폐기물처리', '건축허가표지판', '경계측량및현황측량', '민원처리', '이동식가설화장실반입', '이동식가설화장실반출', '준공청소', '지내력시험', '규준틀설치'] ):
            item.location = '공통가설'
            item.roomname = '공통가설'
            item.floor = '1F'

        #조경공사
        if item.floor.__contains__('조경'):
            item.location = '조경'
            item.roomname = '조경'
            item.floor = '1F'

        # 부대토목공사
        if item.floor.__contains__('부대토목'):
            item.location = '부대토목'
            item.roomname = '부대토목'
            item.floor = '1F'

        # 철거
        if item.floor.__contains__('철거'):
            item.location = '철거'
            item.roomname = '철거'
            item.floor = '1F'

        # 기타공사
        if item.floor.__contains__('기타공사'):
            item.location = '기타공사'
            item.roomname = '기타공사'
            item.floor = '1F'

        # 정화조설치공사
        if item.floor.__contains__('정화조'):
            item.location = '정화조'
            item.roomname = '정화조'
            item.floor = 'FT'



        # 산식 층정리
        # item.formula = item.formula.replace('옥탑','PH').replace('지상','').replace('지하','B').replace('층','F').replace('기초','FT')

        if ((item.floor == '공통가설' or item.floor == '골조가설') and item.formula.__contains__('>') and item.formula.__contains__('<')):
            str = item.formula.split('>')[0].split('<')[1]
            str2 = str.split('F')[0] + 'F'
            str3 = str.split('층')[0] + '층'
            str4 = str.split('붕')[0] + '붕'
            str5 = str.split('초')[0] + '초'
            if (re.match('PH\\d{1,2}F', str) or re.match('B\\d{1,2}F', str) or re.match('\\d{1,2}F', str) or re.match('RF', str) or re.match('PHRF', str)):
                item.location = item.floor
                item.roomname = item.floor
                item.floor = str2
            if (re.match('지상\\d{1,2}층', str) or re.match('지하\\d{1,2}층', str) or re.match('옥탑\\d{1,2}층', str) or re.match('\\d{1}층', str)):
                item.location = item.floor
                item.roomname = item.floor
                item.floor = str3
            if (re.match('지붕', str) or re.match('옥탑지붕', str)):
                item.location = item.floor
                item.roomname = item.floor
                item.floor = str4
            if (re.match('기초', str)):
                item.location = item.floor
                item.roomname = item.floor
                item.floor = str5


        # 층정리
        if (item.type == '외부'
                and (re.match('지상\\d{1,2}층 \\w+', item.floor)
                    or re.match('지하\\d{1,2}층 \\w+', item.floor))):
            split_floor = item.floor.split(' ')
            item.floor = split_floor[0]
            item.location = split_floor[1]
            item.roomname = split_floor[1]


        if (re.match('\\d{1,2}[.] 옥탑\\d{1,2}층', item.floor)):
            item.floor = 'PH' + re.sub(r'[^0-9]', '', item.floor.split(' ')[1]) + 'F'

        if (re.match('\\d{1,2}[.] 지상\\d{1,2}층', item.floor)):
            item.floor = re.sub(r'[^0-9]', '', item.floor.split(' ')[1]) + 'F'

        if (re.match('\\d{1,2}[.] 지하\\d{1,2}층', item.floor)):
            item.floor = 'B' + re.sub(r'[^0-9]', '', item.floor.split(' ')[1]) + 'F'

        if (re.match('옥탑\\d{1,2}층', item.floor)):
            item.floor = 'PH' + re.sub(r'[^0-9]', '', item.floor) + 'F'

        if (re.match('지상\\d{1,2}층', item.floor)):
            item.floor = re.sub(r'[^0-9]', '', item.floor) + 'F'

        if (re.match('지하\\d{1,2}층', item.floor)):
            item.floor = 'B' + re.sub(r'[^0-9]', '', item.floor) + 'F'

        if (re.match('\\d{1,2}층', item.floor)):
            item.floor = re.sub(r'[^0-9]', '', item.floor) + 'F'

        if (item.floor.__contains__('지붕')):
            item.floor = item.floor.replace('지붕', 'RF')

        if (item.floor.__contains__('기초')):
            item.floor = item.floor.replace('기초', 'FT')






    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    for item in items:
        new_sheet.append(item.to_excel())

    sheet = new_workbook.create_sheet(title='집계표')
    sheet.append(['중공종', '품명', '규격', '단위', '수량(할증전)'])
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30
    for item2 in items2:
        sheet.append(item2.to_excel())


    new_workbook.save("C:\\Users\ckddn\Desktop\건축완성.xlsx")

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
