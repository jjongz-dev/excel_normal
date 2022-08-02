import re

from openpyxl import load_workbook, Workbook


from FIN.UMID.ItemStandard import ItemStandard


def excel_normalize(name):
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\건축.xlsx',
        data_only=True)
    # names = excel.get_sheet_names()
    # print(names)
    items = []
    for sheetname in excel.sheetnames:
        worksheet = excel[sheetname]
        location = ""
        for row in worksheet.iter_rows(min_col=0, max_col=8, min_row=4):
            # 호
            if (row[0].value is not None
                    and row[1].value is not None):
                floor = row[0].value.replace(' ','')
                location = row[1].value.split(':')[-1]

            # 행삭제
            if (row[4].value is None):
                continue

            item = ItemStandard(
                floor=floor,
                location=location,
                roomname="roomname",
                name=row[2].value,
                standard=row[3].value,
                unit=row[4].value,
                part='',
                type=sheetname,
                formula=row[5].value,
                sum=row[6].value,
            )
            # print(item.to_excel())
            items.append(item)

    for item in items:
        # 내부호,실 정리
        if (item.type == '내부'):
            item.roomname = item.location

        # TYPE변경
        if (item.type in ['공통가설','가설']):
            item.type = '내부'
            item.floor = '1F'
            item.location = '공통가설'
            item.roomname = '공통가설'

        # 비내력벽
        if (item.type == '비내력벽'):
            item.type = '내부'
            item.floor = ''
            item.location = ''
            item.roomname = ''

        # 계단
        if (item.type == '계단'):
            item.type = '내부'
            item.location = ''
            item.roomname = ''


        # 단위임의체우기
        if (item.unit is None):
            item.unit = 'EA'

        # 창호
        if (item.type == '창호'):
            item.part = item.location.split(',')[0]
            item.floor = ''
            item.location = ''
            item.roomname = ''

        # 외부
        if (item.type == '외부'):
            item.location = ''
            item.roomname = ''

        # 층정리
        if (re.match('\\d{1,2}층', item.floor)):
            item.floor = re.sub(r'[^0-9]', '', item.floor).lstrip("0") + 'F'

        if (re.match('B\\d{1,2}층', item.floor)):
            item.floor = 'B' + re.sub(r'[^0-9]', '', item.floor).lstrip("0") + 'F'

        if (re.match('P\\d{1,2}층', item.floor)):
            item.floor = 'PH' + re.sub(r'[^0-9]', '', item.floor).lstrip("0") + 'F'

        if (item.floor == 'FTPIT층'):
            item.floor = 'B2F'



    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    for item in items:
        new_sheet.append(item.to_excel())

    new_workbook.save("C:\\Users\ckddn\Desktop\건축완성.xlsx")

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
