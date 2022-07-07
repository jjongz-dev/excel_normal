# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.
import dataclasses
import re

from openpyxl import load_workbook, Workbook
from dataclasses import dataclass

from openpyxl.cell import MergedCell

from QuantityItemStandard import QuantityItemStandard
from QuantityItemStandard2 import QuantityItemStandard2


def excel_normalize(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.
    workbook = load_workbook(
        'C:\\전기.xlsm',
        data_only=True)
    # names = workbook.get_sheet_names()
    # print(names)


    worksheet = workbook['목록별산출서']
    items = []
    category = ""
    for row in worksheet.iter_rows(min_col=4, max_col=13, min_row=1):
        for cell in row:
            if isinstance(cell, MergedCell):
                if (row[0].value is None):
                    continue
                category = ''.join(re.compile('[가-힣]').findall(row[0].value.split('::')[1]))
                break
        if row[4].value is None:
            continue
        if row[4].value == '명칭':
            continue

        item = QuantityItemStandard(
            category = category,
            name=row[4].value,
            standard=row[5].value,
            unit=row[6].value,
            formula=row[1].value,
            unit_formula=row[7].value,
            sum=row[9].value,
        )
        if item.formula is None:
            item.formula = items[-1].formula
        print(item)
        items.append(item)
    print(items.__sizeof__())






    worksheet2 = workbook['산출집계']
    items2 = []
    for row in worksheet2.iter_rows(min_col=4, max_col=10, min_row=1):
        for cell in row:
            if isinstance(cell, MergedCell):
                if (row[0].value is None):
                    continue
                category = ''.join(re.compile('[가-힣]').findall(row[0].value.split('::')[1]))
                break
        if row[4].value is None:
            continue
        if row[6].value == '소 계':
            continue
        item2 = QuantityItemStandard2(
            category = category,
            name=row[0].value,
            standard=row[1].value,
            unit=row[2].value,
            sum=row[6].value,
        )
        items2.append(item2)
    print(items2)



    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '전기(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    for item in items:
        new_sheet.append(item.to_excel())

    sheet = new_workbook.create_sheet(title='집계표')
    sheet.append(['중공종', '품명', '규격', '단위', '수량(할증전)'])
    for item2 in items2:
        sheet.append(item2.to_excel())

    workbook = load_workbook(
        'C:\\통신.xlsm',
        data_only=True)
    # names = workbook.get_sheet_names()
    # print(names)
    worksheet = workbook['목록별산출서']
    items = []
    for row in worksheet.iter_rows(min_col=4, max_col=13, min_row=1):
        for cell in row:
            if isinstance(cell, MergedCell):
                if (row[0].value is None):
                    continue
                category = ''.join(re.compile('[A-Za-z가-힣]').findall(row[0].value.split('::')[1]))
                break
        if row[4].value is None:
            continue
        if row[4].value == '명칭':
            continue
        item = QuantityItemStandard(
            category=category,
            name=row[4].value,
            standard=row[5].value,
            unit=row[6].value,
            formula=row[1].value,
            unit_formula=row[7].value,
            sum=row[9].value,
        )
        if item.formula is None:
            item.formula = items[-1].formula
        items.append(item)
    print(items.__sizeof__())

    worksheet2 = workbook['산출집계']
    items2 = []
    for row in worksheet2.iter_rows(min_col=4, max_col=10, min_row=1):
        for cell in row:
            if isinstance(cell, MergedCell):
                if (row[0].value is None):
                    continue
                category = ''.join(re.compile('[A-Za-z가-힣]').findall(row[0].value.split('::')[1]))
                break
        if row[4].value is None:
            continue
        if row[6].value == '소 계':
            continue
        item2 = QuantityItemStandard2(
            category=category,
            name=row[0].value,
            standard=row[1].value,
            unit=row[2].value,
            sum=row[6].value,
        )
        items2.append(item2)
    print(items2)

    for item in items:
        new_sheet.append(item.to_excel())

    for item2 in items2:
        sheet.append(item2.to_excel())

    workbook = load_workbook(
        'C:\\소방.xlsm',
        data_only=True)
    # names = workbook.get_sheet_names()
    # print(names)
    worksheet = workbook['목록별산출서']
    items = []
    for row in worksheet.iter_rows(min_col=4, max_col=13, min_row=1):
        for cell in row:
            if isinstance(cell, MergedCell):
                if (row[0].value is None):
                    continue
                category = ''.join(re.compile('[가-힣]').findall(row[0].value.split('::')[1]))
                break
        if row[4].value is None:
            continue
        if row[4].value == '명칭':
            continue
        item = QuantityItemStandard(
            category= category,
            name=row[4].value,
            standard=row[5].value,
            unit=row[6].value,
            formula=row[1].value,
            unit_formula=row[7].value,
            sum=row[9].value,
        )
        if item.formula is None:
            item.formula = items[-1].formula
        items.append(item)
    print(items.__sizeof__())

    worksheet2 = workbook['산출집계']
    items2 = []
    for row in worksheet2.iter_rows(min_col=4, max_col=10, min_row=1):
        for cell in row:
            if isinstance(cell, MergedCell):
                if (row[0].value is None):
                    continue
                category = ''.join(re.compile('[가-힣]').findall(row[0].value.split('::')[1]))
                break
        if row[4].value is None:
            continue
        if row[6].value == '소 계':
            continue
        item2 = QuantityItemStandard2(
            category=category,
            name=row[0].value,
            standard=row[1].value,
            unit=row[2].value,
            sum=row[6].value,
        )
        items2.append(item2)
    print(items2)

    for item in items:
        new_sheet.append(item.to_excel())

    for item2 in items2:
        sheet.append(item2.to_excel())

    new_workbook.save("C:\\Users\\Box\\Desktop\\외부적산\\test.xlsx")


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
