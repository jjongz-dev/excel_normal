from openpyxl import load_workbook
from openpyxl.cell import MergedCell


def excel_normalize(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.
    workbook = load_workbook(
        'C:\\서울시 은평구 역촌동 77-9,47 근생주택 신축공사-전기(산출).xlsm',
        data_only=True)
    # names = workbook.get_sheet_names()
    # print(names)
    worksheet = workbook['목록별산출서']
    for row in worksheet.iter_rows(max_row=5, min_col=4):
        for cell in row:
            if isinstance(cell, MergedCell):
                if ( row[0].value is None):
                    continue
                print(row[0].value)
                break

    # print(worksheet["D3"].value)
if __name__ == '__main__':
    excel_normalize('PyCharm')