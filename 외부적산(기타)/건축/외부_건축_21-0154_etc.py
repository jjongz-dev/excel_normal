from openpyxl import load_workbook, Workbook
from ExcelStandard import ExcelStandard
from ExcelGroup import ExcelGroup
from collections import defaultdict
from pprint import pprint as pprint
from datetime import datetime

import re
import platform
import subprocess
import ReplaceFinEarthWork
import ReplaceFinDuplicateDelete
import ReplacePersonal


fileCreateDate = datetime.strftime(datetime.today(), '%y%m%d_%H%M')
systemOs = platform.system()


# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '21-0154_etc'
##################################
if systemOs == 'Darwin':
    openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/건축.xlsx'
    saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/end-1-kun-' + fileCreateDate + '.xlsx'
else:
    openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축.xlsx'
    saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)


    내역서목록 = []
    산출서목록 = []
    파싱시작점기준문자 = ['부위', '도형', '구분', '코드', '품목', '품명']


    파싱시트목록 = ['1. 본동-물량산출서']

    print('=================================')
    for 시트명 in 파싱시트목록:
        if 시트명 in excel.sheetnames:
            print("파싱시트체크 : ", 시트명, '(O)')
        else:
            print("파싱시트체크 : ", 시트명, '(X) - 확인필요')
    print('=================================')



    for sheetname in 파싱시트목록:

        if sheetname in excel.sheetnames:
            worksheet = excel[sheetname]

            층확정 = ''
            호확정 = ''
            실확정 = ''
            품명확정 = ''
            규격확정 = ''
            단위확정 = ''
            부위확정 = ''
            타입확정 = ''
            산식확정 = ''
            수량확정 = ''

            for row in worksheet.iter_rows(min_row=2):

                위치값 = row[0].value
                품명값 = row[1].value
                규격값 = row[2].value
                단위값 = row[3].value
                부위값 = row[4].value
                산식값 = row[5].value
                수량값 = row[8].value
                개소값 = row[7].value

    ####### 타입 / 층 / 실명 뽑아내기  위치에만 데이터가 있는 항목들 대상  S ##############################

                if (위치값 is not None
                        and (품명값 is None or 품명값 == '')
                        and (규격값 is None or 규격값 == '')
                        and (단위값 is None or 단위값 == '')
                        and (부위값 is None or 부위값 == '')
                        and (산식값 is None or 산식값 == '')
                        and (수량값 is None or 수량값 == '')
                ):

                    if '>' in 위치값:
                        타입확정 = 위치값.split('>')[-1].strip()
                        continue

                    if "/" in 위치값:

                        위치값_분리_슬래쉬 = 위치값.split('/')
                        창호명 = 위치값_분리_슬래쉬[0].strip()
                        창호사이즈 = 위치값_분리_슬래쉬[1].split('(')[0]
                        창호가로사이즈 = 창호사이즈.split('*')[0]
                        창호세로사이즈 = 창호사이즈.split('*')[1]

                        층확정 = 창호명
                        호확정 = 창호명
                        실확정 = 창호명

                        #print(위치, '  ==> ', 위치값_분리_슬래쉬[0], '/', 위치값_분리_슬래쉬[1], ' ===>', 호확정 ,' ===>', 창호가로사이즈, "/", 창호세로사이즈)

                        continue

                    if '[문]' in 위치값 or '[창]' in 위치값:
                        continue

                    #print(위치값, '품명값 : ', 품명값)
                    if ('[가로]X' in 위치값 or '[둘레]L' in 위치값 or '[세로]Y' in 위치값) and (품명값 is None or 품명값 == '') :
                        continue

                    #B301 근린생활, 전기실  등에서  층과 실 추출

                    위치값_분리_빈칸 = 위치값.split(' ')
                    위치값_분리_빈칸_길이 = len(위치값_분리_빈칸)

                    if 위치값_분리_빈칸_길이 == 1:
                        층확정 = 위치값_분리_빈칸[0]
                        호확정 = 위치값_분리_빈칸[0]
                        실확정 = 타입확정
                        continue
                    else:
                        층확정 = 위치값_분리_빈칸[0]
                        호확정 = 위치값_분리_빈칸[1]
                        실확정 = 위치값_분리_빈칸[1]
                        continue

                if 위치값 == '위치' and 품명값 == '품명' and 규격값 == '규격':
                    continue

                if (위치값 is None or 위치값 == '') and (품명값 is None or 품명값 == '') and (규격값 is None or 규격값 == ''):
                    continue

                if ('[가로]X' in 위치값 or '[둘레]L' in 위치값 or '공사명' in 위치값) and (품명값 is None or 품명값 == '') and (규격값 is None or 규격값 == ''):
                    continue


                층확정원본값 = 층확정

                층별데이터복사 = False

                층목록 = []

                if 층확정 is not None:
                    # 2~5F
                    지상층패턴 = r'(\d)[~](\d)[층]'
                    지상층검색결과 = re.search(지상층패턴, 층확정)

                if 지상층검색결과 is not None:

                    층별데이터복사 = True

                    시작층 = int(지상층검색결과.group(1))
                    종료층 = int(지상층검색결과.group(2)) + 1

                    if 시작층 >= 종료층:
                        시작층 = int(지상층검색결과.group(2))
                        종료층 = int(지상층검색결과.group(1)) + 1

                    for 층 in range(시작층, 종료층):
                        층값 = f'{층}F'
                        층목록.append(층값)


                if 품명값 is not None and 수량값 is not None:

                    품명확정 = 품명값
                    규격확정 = 규격값
                    단위확정 = 단위값
                    부위확정 = 부위값
                    산식확정 = 산식값
                    수량확정 = 수량값
                    개소확정 = 개소값


                    if 층별데이터복사:

                        원본수량 = 수량확정
                        분할수량 = 0
                        분할수량합계 = 0
                        총층수 = len(층목록)

                        for index, 층 in enumerate(층목록):

                            층변환확정 = 층
                            분할수량 = round(수량확정/총층수, 3)

                            if index == 총층수-1:
                                분할수량 = 원본수량 - 분할수량합계
                            else:
                                분할수량합계 = 분할수량 + 분할수량합계

                            내역 = ExcelStandard(
                                층=층변환확정,
                                호=호확정,
                                실=실확정,
                                대공종='건축',
                                중공종='',
                                코드='',
                                품명=품명확정,
                                규격=규격확정,
                                단위=단위확정,
                                부위=부위확정,
                                타입=타입확정,
                                산식=산식확정,
                                수량=분할수량,
                                Remark='',
                                개소=개소확정
                            )
                            산출서목록.append(내역)
                    else:

                        층확정 = 층확정원본값

                        내역 = ExcelStandard(
                            층=층확정,
                            호=호확정,
                            실=실확정,
                            대공종='건축',
                            중공종='',
                            코드='',
                            품명=품명확정,
                            규격=규격확정,
                            단위=단위확정,
                            부위=부위확정,
                            타입=타입확정,
                            산식=산식확정,
                            수량=수량확정,
                            Remark='',
                            개소=개소확정
                        )
                        산출서목록.append(내역)


    # 품명 규격 개인별 지정 변경 S #######################
    for 내역 in 산출서목록:
        ReplacePersonal.launch(내역)
    # 품명 규격 개인별 지정 변경 E #######################



    # 공종별내역서 ###################################################################################

    if '공종별내역서' in excel.sheetnames:

        worksheet = excel['공종별내역서']
        내역서시작줄 = 0
        중공종확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        수량확정 = ''

        for 가로줄번호, row in enumerate(worksheet.rows):

            if row[0].value.replace(' ', '') in 파싱시작점기준문자:
                내역서시작줄 = 가로줄번호 + 3
                break

        for row in worksheet.iter_rows(min_row=내역서시작줄):

            품명값 = row[0].value
            규격값 = row[1].value
            단위값 = row[2].value
            수량값 = row[3].value

            if 품명값 is not None and (단위값 is None or 단위값 == ''):
                if '합  계' in 품명값:
                    continue
                else:
                    중공종 = 품명값.replace(' ', '')
                    중공종확정 = 중공종.split('.')[-1]
                    continue

            if 품명값 is not None and (단위값 is not None or 단위값 != '') and (수량값 is not None and 수량값 != 0):

                품명확정 = 품명값
                규격확정 = 규격값
                단위확정 = 단위값
                수량확정 = 수량값

                내역 = ExcelGroup(
                    중공종=중공종확정,
                    품명=품명확정,
                    규격=규격확정,
                    단위=단위확정,
                    수량=수량확정
                )
                내역서목록.append(내역)


    # 엑셀 처리 완료 ###################################################################################




    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark', '개소(확인용)']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30

    for 내역 in 산출서목록:
       new_sheet.append(내역.to_excel())

    sheet = new_workbook.create_sheet(title='집계표')
    sheet.append(['중공종', '품명', '규격', '단위', '수량(할증전)'])
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30

    for 내역 in 내역서목록:
        sheet.append(내역.to_excelGroup())

    new_workbook.save(saveFilePath)



    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])
    elif systemOs == "Windows":
        subprocess.Popen(saveFilePath, shell=True)


if __name__ == '__main__':
    excel_normalize('PyCharm')
