from openpyxl import load_workbook, Workbook
from ExcelStandard import ExcelStandard
from datetime import datetime
from collections import defaultdict
from pprint import pprint as pprint
import platform
import subprocess
import ReplacePersonal

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')

# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0323_etc'
##################################

openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/건축.xlsx'
saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/건축완성-' + fileCreateDate + '.xlsx'

# openFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\건축.xlsx'
# saveFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\건축완성-' + fileCreateDate + '.xlsx'

# openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축.xlsx'
# saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)

    내역목록 = []

    if 'Table 1' in excel.sheetnames:
        worksheet = excel['Table 1']



        층확정 = ''
        호확정 = ''
        실확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        부위확정 = ''
        타입확정 = ''
        산식확정 = ''
        수량확정 = ''

        품규시작 = False

        for row in worksheet.iter_rows(min_row=2):

            실값 = row[0].value
            구분값 = row[1].value
            층값 = row[2].value
            산식값 = row[2].value
            수량값 = row[3].value

            if 실값 is not None and 'A-' in 실값:

                호확정 = f'{실값}{구분값}'
                품명목록 = defaultdict(str)
                규격목록 = defaultdict(str)

                층확정 = 층값.replace('평면도', '').replace('평면', '')
                print(실값, 구분값, 층확정)
                continue

            if 호확정 == 'A-16':


                if 수량값 is not None and 실값 != '실 별':

                    print(실값, 구분값, 수량값)
                    내역 = ExcelStandard(
                        층=층확정,
                        호=호확정,
                        실='',
                        대공종='건축',
                        중공종='',
                        코드='',
                        품명=실값,
                        규격=구분값,
                        단위='',
                        부위='',
                        타입='',
                        산식=산식값,
                        수량=수량값,
                        Remark='',
                        개소=''
                    )
                    내역목록.append(내역)


            if 실값 is not None and '실 별' in 실값:

                품규시작 = True
                for num in range(4, 16):
                    품명값 = row[num].value
                    if 품명값 is not None:
                        품명목록[num] = 품명값
                continue

            if 품규시작:
                품규시작 =False
                for num in range(4, 16):
                    규격값 = row[num].value
                    if 규격값 is not None:
                        규격목록[num] = 규격값
                continue


            if 품규시작 is False and 실값 is None and 구분값 is None:
                continue

            if 실값 == '계':
                continue

            if 실값 is not None:
                실확정 = 실값

            for num in range(4, 16):



                품명값 = 품명목록[num]
                규격값 = 규격목록[num]
                수량확정 = row[num].value

                if 품명값 is not None:
                    품명확정 = 품명값

                if 규격값 is not None:
                    규격확정 = 규격값

                if 수량확정 is not None:
                    산식확정= 산식값

                    #print(층확정, 실확정, 구분값, 품명확정, 규격확정, 수량확정)

                    내역 = ExcelStandard(
                        층=층확정,
                        호=호확정,
                        실=실확정,
                        대공종='건축',
                        중공종='',
                        코드='',
                        품명=품명확정,
                        규격=규격확정,
                        단위='',
                        부위='',
                        타입='',
                        산식=산식확정,
                        수량=수량확정,
                        Remark='',
                        개소=''
                    )
                    내역목록.append(내역)


            continue



    # 품명 규격 개인별 지정 변경 S #######################
    for 내역 in 내역목록:
        ReplacePersonal.launch(내역)
    # 품명 규격 개인별 지정 변경 E #######################



    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축완성'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark', '개소(확인용)']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 15
    new_sheet.column_dimensions["H"].width = 15

    for 내역 in 내역목록:
       new_sheet.append(내역.to_excel())

    new_workbook.save(saveFilePath)

    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])


if __name__ == '__main__':
    excel_normalize('PyCharm')

