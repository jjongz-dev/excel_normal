from openpyxl import load_workbook, Workbook
from ExcelStandard import ExcelStandard
from ExcelGroup import ExcelGroup
from datetime import datetime

import re
import platform
import subprocess
import ReplacePersonal



from Architect.RC.UMID.ItemStandard import ItemStandard


fileCreateDate = datetime.strftime(datetime.today(), '%y%m%d_%H%M')
systemOs = platform.system()


# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '22-0669_umid'
##################################


if systemOs == 'Darwin':
    openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/구조.xlsx'
    saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/구조완성-' + fileCreateDate + '.xlsx'
else:
    openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조.xlsx'
    saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)
    items = []
    for sheetname in excel.sheetnames:
        worksheet = excel[sheetname]
        location = ""
        part = ""
        for row in worksheet.iter_rows(min_col=0, max_col=8, min_row=4):
            # 호
            if (row[0].value is not None
                    and row[1].value is not None):
                location = row[1].value.split(':')[-1]

            # 행삭제
            if (row[4].value is None):
                continue

            # 부위
            if (row[1].value is not None
                    and row[4].value is not None):
                part = row[1].value

            item = ItemStandard(
                floor=sheetname,
                location=location,
                name=row[2].value,
                standard=row[3].value,
                part=part,
                formula=row[5].value,
                sum=row[6].value,
            )
            # print(item.to_excel())
            items.append(item)

    for item in items:
        # 층정리
        if (re.match('\\d{1,2}', item.floor)):
            item.floor = re.sub(r'[^0-9]', '', item.floor).lstrip("0") + 'F'

        if (re.match('B\\d{1,2}', item.floor)):
            item.floor = 'B' + re.sub(r'[^0-9]', '', item.floor).lstrip("0") + 'F'

        if (re.match('P\\d{1,2}', item.floor)):
            item.floor = 'PH' + re.sub(r'[^0-9]', '', item.floor).lstrip("0") + 'F'

        if (item.floor == 'FTPIT'):
            item.floor = 'B2F'

    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 15
    new_sheet.column_dimensions["H"].width = 15

    for 내역 in items:
        new_sheet.append(내역.to_excel())

    new_workbook.save(saveFilePath)


    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])
    elif systemOs == "Windows":
        subprocess.Popen(saveFilePath, shell=True)


if __name__ == '__main__':
    excel_normalize('PyCharm')
