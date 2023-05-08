import re

from openpyxl import load_workbook, Workbook
from ExcelStandard import ExcelStandard
from datetime import datetime
import platform
import subprocess
import ReplacePersonal

fileCreateDate = datetime.strftime(datetime.today(), '%y%m%d_%H%M')
systemOs = platform.system()

# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '21-0154_etc'
##################################


if systemOs == 'Darwin':
    openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/구조.xlsx'
    saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/end-2-gujo-' + fileCreateDate + '.xlsx'
else:
    openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조.xlsx'
    saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)

    산출서목록 = []

    파싱시트목록 = ['본관동-물량산출서']
    print('=================================')
    for 시트명 in 파싱시트목록:
        if 시트명 in excel.sheetnames:
            print("파싱시트체크 : ", 시트명, '(O)')
        else:
            print("파싱시트체크 : ", 시트명, '(X) - 확인필요')
    print('=================================')



    for sheetname in 파싱시트목록:

        if sheetname in excel.sheetnames:
            worksheet = excel[sheetname]


            층확정 = ''
            호확정 = ''
            실확정 = ''
            품명확정 = ''
            규격확정 = ''
            부위확정 = ''
            산식확정 = ''
            수량확정 = ''

            for 가로줄번호, row in enumerate(worksheet.rows):

                엑셀가로줄번호 = 가로줄번호 + 1

                층값 = row[0].value
                개소값 = row[2].value

                콘크리트종류값 = row[4].value
                콘크리트산출식값 = row[5].value
                콘크리트소계값 = row[10].value

                거푸집종류값 = row[11].value
                거푸집산출식값 = row[13].value
                거푸집소계값 = row[18].value

                철근규격값 = row[21].value
                철근산출식값 = row[23].value
                철근소계값 = row[30].value

                if 엑셀가로줄번호 ==1:
                    continue

                if 층값 is not None:

                    층값 = 층값.replace('\n', '').replace('\r', '')

                # for 세로줄번호, col in enumerate(list(row)):
                #     print(층값, 세로줄번호, col)

                if 콘크리트종류값 in ['콘크리트(M3)', '종류', '계', '물 량 산 출 서']:
                    continue

                if 층값 in ['물 량 산 출 서', '계']:
                    continue

                if row[17].value in ['D', 'H D', 'H D/s', 'SHD', 'SHD/s', 'UHD', 'SSH']:
                    continue

                if 층값 is not None and '본관동' in 층값:
                    호확정 = 층값.split(' ')[-2]
                    continue

                if 층값 is not None and 개소값 is not None:
                    다음줄층값 = worksheet[엑셀가로줄번호 + 1][0].value
                    if 다음줄층값 is not None:
                        층확정 = 다음줄층값
                        부위확정 = 층값

                층확정원본값 = 층확정

                층별데이터복사 = False

                층목록 = []

                if 층확정 is not None:

                    # B2.B1, B2~B1, B2B1
                    지하층패턴 = r'[B](\d+)[B](\d+)'
                    지하층검색결과 = re.search(지하층패턴, 층확정)

                    # N2.N4, N2~N4, N2N4
                    지상층패턴 = r'[N](\d+)[N](\d+)'
                    지상층검색결과 = re.search(지상층패턴, 층확정)

                    # B2.N4, B2~N4, B2N4
                    혼합층패턴 = r'[B](\d+)[N](\d+)'
                    혼합층검색결과 = re.search(혼합층패턴, 층확정)

                    # NO4.5.6
                    지상층패턴1 = r'[NO](\d)[.]'
                    지상층패턴1검색결과 = re.search(지상층패턴1, 층확정)

                    if 지하층검색결과 is not None:

                        층별데이터복사 = True

                        시작층 = int(지하층검색결과.group(1))
                        종료층 = int(지하층검색결과.group(2)) + 1

                        if 시작층 >= 종료층:
                            시작층 = int(지하층검색결과.group(2))
                            종료층 = int(지하층검색결과.group(1)) + 1

                        for 층 in range(시작층, 종료층):
                            층값 = f'B{str(층).zfill(2)}'
                            층목록.append(층값)

                    if 지상층검색결과 is not None:

                        층별데이터복사 = True

                        시작층 = int(지상층검색결과.group(1))
                        종료층 = int(지상층검색결과.group(2)) + 1

                        if 시작층 >= 종료층:
                            시작층 = int(지상층검색결과.group(2))
                            종료층 = int(지상층검색결과.group(1)) + 1

                        for 층 in range(시작층, 종료층):
                            층값 = f'N{str(층).zfill(2)}'
                            층목록.append(층값)

                    if 혼합층검색결과 is not None:

                        층별데이터복사 = True

                        지하층층수 = int(혼합층검색결과.group(1)) + 1

                        for 층 in range(1, 지하층층수):
                            층값 = f'B{str(층).zfill(2)}'
                            층목록.append(층값)

                        지상층층수 = int(혼합층검색결과.group(2)) + 1

                        for 층 in range(1, 지상층층수):
                            층값 = f'N{str(층).zfill(2)}'
                            층목록.append(층값)

                    if 지상층패턴1검색결과 is not None:

                        층별데이터복사 = True

                        층정보추출 = re.sub(r'[^\d.]', '', 층확정)
                        층정보추출분리_점 = 층정보추출.split('.')
                        for 층 in 층정보추출분리_점:
                            층값 = f'NO{str(층).zfill(2)}'
                            층목록.append(층값)


                # 콘크리트 산출 Start ##############################################################

                if 콘크리트종류값 is not None and 콘크리트산출식값 is not None:

                    다음줄층값 = worksheet[엑셀가로줄번호 + 1][0].value

                    if 콘크리트소계값 is None or 콘크리트소계값 == '':
                        다음줄콘크리트산출식값 = str(worksheet[엑셀가로줄번호 + 1][5].value)
                        다음줄콘크리트소계값 = worksheet[엑셀가로줄번호 + 1][10].value
                        다다음줄콘크리트산출식값 = str(worksheet[엑셀가로줄번호 + 2][5].value)
                        다다음줄콘크리트소계값 = worksheet[엑셀가로줄번호 + 2][10].value

                        if 다음줄콘크리트소계값 is not None:
                            콘크리트산출식값 = f'{콘크리트산출식값}{다음줄콘크리트산출식값}'
                            콘크리트소계값 = 다음줄콘크리트소계값
                        elif 다다음줄콘크리트소계값 is not None:
                            콘크리트산출식값 = f'{콘크리트산출식값}{다음줄콘크리트산출식값}{다다음줄콘크리트산출식값}'
                            콘크리트소계값 = 다다음줄콘크리트소계값

                    규격확정 = 콘크리트종류값
                    산식확정 = 콘크리트산출식값
                    수량확정 = 콘크리트소계값

                    #print(엑셀가로줄번호, 층값, 층확정, 호확정, 규격확정, 부위확정, 산식확정, 수량확정, '/', 다음줄콘크리트산출식값)

                    if 층별데이터복사:

                        원본수량 = 수량확정
                        분할수량 = 0
                        분할수량합계 = 0
                        총층수 = len(층목록)

                        for index, 층 in enumerate(층목록):

                            층변환 = str(층).zfill(2)
                            층변환확정 = 층

                            산식변환확정 = 산식확정[:산식확정.rfind('*')]
                            분할수량 = round(수량확정/총층수, 3)

                            if index == 총층수 - 1:
                                분할수량 = 원본수량 - 분할수량합계
                            else:
                                분할수량합계 = 분할수량 + 분할수량합계

                            내역 = ExcelStandard(
                                층=층변환확정,
                                호=호확정,
                                실='',
                                대공종='건축',
                                중공종='철근콘크리트공사',
                                코드='',
                                품명='콘크리트',
                                규격=규격확정,
                                단위='',
                                부위=부위확정,
                                타입='구조',
                                산식=산식변환확정,
                                수량=분할수량,
                                Remark='',
                                개소=실확정
                            )
                            산출서목록.append(내역)

                    else:

                        층확정 = 층확정원본값

                        내역 = ExcelStandard(
                            층=층확정,
                            호=호확정,
                            실='',
                            대공종='건축',
                            중공종='철근콘크리트공사',
                            코드='',
                            품명='콘크리트',
                            규격=규격확정,
                            단위='',
                            부위=부위확정,
                            타입='구조',
                            산식=산식확정,
                            수량=수량확정,
                            Remark='',
                            개소=''
                        )
                        산출서목록.append(내역)

                # 콘크리트 산출 End ##############################################################

                # 거푸집 산출 Start ##############################################################

                if 거푸집종류값 is not None and 거푸집산출식값 is not None:

                    if 거푸집소계값 is None or 거푸집소계값 == '':
                        다음줄거푸집산출식값 = str(worksheet[엑셀가로줄번호 + 1][13].value)
                        다음줄거푸집소계값 = worksheet[엑셀가로줄번호 + 1][18].value
                        다다음줄거푸집산출식값 = str(worksheet[엑셀가로줄번호 + 2][13].value)
                        다다음줄거푸집소계값 = worksheet[엑셀가로줄번호 + 2][18].value

                        if 다음줄거푸집소계값 is not None:
                            거푸집산출식값 = f'{거푸집산출식값}{다음줄거푸집산출식값}'
                            거푸집소계값 = 다음줄거푸집소계값
                        elif 다다음줄거푸집소계값 is not None:
                            거푸집산출식값 = f'{거푸집산출식값}{다음줄거푸집산출식값}{다다음줄거푸집산출식값}'
                            거푸집소계값 = 다다음줄거푸집소계값

                    규격확정 = 거푸집종류값
                    산식확정 = 거푸집산출식값
                    수량확정 = 거푸집소계값

                    #print(엑셀가로줄번호, 층값, 층확정, 호확정, 규격확정, 부위확정, 산식확정, 수량확정, '/', 다음줄거푸집산출식값)

                    if 층별데이터복사:

                        원본수량 = 수량확정
                        분할수량 = 0
                        분할수량합계 = 0
                        총층수 = len(층목록)

                        for index, 층 in enumerate(층목록):

                            층변환 = str(층).zfill(2)
                            층변환확정 = 층

                            산식변환확정 = 산식확정[:산식확정.rfind('*')]
                            분할수량 = round(수량확정/총층수, 3)

                            if index == 총층수 - 1:
                                분할수량 = 원본수량 - 분할수량합계
                            else:
                                분할수량합계 = 분할수량 + 분할수량합계

                            내역 = ExcelStandard(
                                층=층변환확정,
                                호=호확정,
                                실='',
                                대공종='건축',
                                중공종='철근콘크리트공사',
                                코드='',
                                품명='거푸집',
                                규격=규격확정,
                                단위='',
                                부위=부위확정,
                                타입='구조',
                                산식=산식변환확정,
                                수량=분할수량,
                                Remark='',
                                개소=실확정
                            )
                            산출서목록.append(내역)
                    else:

                        층확정 = 층확정원본값

                        내역 = ExcelStandard(
                            층=층확정,
                            호=호확정,
                            실='',
                            대공종='건축',
                            중공종='철근콘크리트공사',
                            코드='',
                            품명='거푸집',
                            규격=규격확정,
                            단위='',
                            부위=부위확정,
                            타입='구조',
                            산식=산식확정,
                            수량=수량확정,
                            Remark='',
                            개소=''
                        )
                        산출서목록.append(내역)

                # 거푸집 산출 End ##############################################################

                # 철근 산출 Start ##############################################################

                if 철근규격값 is not None and 철근산출식값 is not None:

                    if 철근소계값 is None or 철근소계값 == '':
                        다음줄철근산출식값 = str(worksheet[엑셀가로줄번호 + 1][23].value)
                        다음줄철근소계값 = worksheet[엑셀가로줄번호 + 1][30].value
                        다다음줄철근산출식값 = str(worksheet[엑셀가로줄번호 + 2][23].value)
                        다다음줄철근소계값 = worksheet[엑셀가로줄번호 + 2][30].value

                        if 다음줄철근소계값 is not None:
                            철근산출식값 = f'{철근산출식값}{다음줄철근산출식값}'
                            철근소계값 = 다음줄철근소계값
                        elif 다다음줄철근소계값 is not None:
                            철근산출식값 = f'{철근산출식값}{다음줄철근산출식값}{다다음줄철근산출식값}'
                            철근소계값 = 다다음줄철근소계값

                    규격확정 = 철근규격값
                    산식확정 = 철근산출식값
                    수량확정 = 철근소계값

                    #print(엑셀가로줄번호, 층값, 층확정, 호확정, 규격확정, 부위확정, 산식확정, 수량확정)
                    if 층별데이터복사:

                        원본수량 = 수량확정
                        분할수량 = 0
                        분할수량합계 = 0
                        총층수 = len(층목록)

                        for index, 층 in enumerate(층목록):

                            층변환 = str(층).zfill(2)
                            층변환확정 = 층

                            산식변환확정 = 산식확정[:산식확정.rfind('*')]
                            분할수량 = round(수량확정/총층수, 3)

                            if index == 총층수-1:
                                분할수량 = 원본수량 - 분할수량합계
                            else:
                                분할수량합계 = 분할수량 + 분할수량합계

                            내역 = ExcelStandard(
                                층=층변환확정,
                                호=호확정,
                                실='',
                                대공종='건축',
                                중공종='철근콘크리트공사',
                                코드='',
                                품명='철근',
                                규격=규격확정,
                                단위='',
                                부위=부위확정,
                                타입='구조',
                                산식= 산식변환확정,
                                수량=분할수량,
                                Remark='',
                                개소=실확정
                            )
                            산출서목록.append(내역)
                    else:

                        층확정 = 층확정원본값

                        내역 = ExcelStandard(
                            층=층확정,
                            호=호확정,
                            실='',
                            대공종='건축',
                            중공종='철근콘크리트공사',
                            코드='',
                            품명='철근',
                            규격=규격확정,
                            단위='',
                            부위=부위확정,
                            타입='구조',
                            산식=산식확정,
                            수량=수량확정,
                            Remark='',
                            개소=''
                        )
                        산출서목록.append(내역)


                # 철근 산출 End ##############################################################


    # 품명 규격 개인별 지정 변경 S #######################
    for 내역 in 산출서목록:
        ReplacePersonal.launch(내역)
    # 품명 규격 개인별 지정 변경 E #######################




    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조완성'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark','개소']
    new_sheet.append(head_title)
    # new_sheet.column_dimensions["G"].width = 15
    # new_sheet.column_dimensions["H"].width = 15
    new_sheet.column_dimensions["L"].width = 80

    for 내역 in 산출서목록:
       new_sheet.append(내역.to_excel())

    new_workbook.save(saveFilePath)


    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])
    elif systemOs == "Windows":
        subprocess.Popen(saveFilePath, shell=True)


if __name__ == '__main__':
    excel_normalize('PyCharm')
