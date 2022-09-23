from openpyxl import load_workbook, Workbook

from Architect.GoogleSheet.ItemStandard import ItemStandard



def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\창호.xlsx',
        data_only=True)

    items = []
    worksheet = excel['초안']
    for row in worksheet.iter_rows(min_col=0, max_col=25, min_row=3):
        if ( row[0].value is None
                and row[1].value is None
                and row[2].value is None
                and row[3].value is None
                and row[4].value is None
        ):
            continue

        # 이름변경 temp_windows_name
        try:
            if ( row[0].value is not None
            ):
                temp_windows_name = f"{row[0].value} | {float(row[1].value)}*{float(row[2].value)} | M2"
                if ( row[15].value is not None
                ):
                    temp_windows_name = f"{row[0].value}(시스템도어포함) | {float(row[1].value)}*{float(row[2].value)} | M2"
        except:
            print('창호이름오류')


        # 유리변경 temp_glass_standard
        try:
            if ( row[3].value is None
            ):
                temp_standard = ''

            if ( row[3].value is not None
                    and row[4].value is not None
                    and row[5].value is not None
            ):
                temp_glass_standard =f"{row[3].value}유리 | {row[4].value},{row[5].value}mm | M2"
                if ( row[6].value is not None):
                    temp_glass_standard = f"{row[3].value}유리(아르곤) | {row[4].value},{row[5].value}mm | M2"
            else:
                temp_glass_standard = ''
        except:
            print('유리오류')

        # 소방진입창 fire_entrance
        try:
            if ( row[14].value is not None
            ):
                temp_fire_entrance = f"{row[3].value}유리 | {row[4].value},{row[5].value}mm,파인크러쉬 | M2"
                if (row[6].value is not None):
                    temp_fire_entrance = f"{row[3].value}유리(아르곤) | {row[4].value},{row[5].value}mm,파인크러쉬 | M2"
            else:
                temp_fire_entrance = ''
        except:
            print('소방진입창오류')

        # 도어변경 temp_glass_door
        try:
            if ( row[7].value is None
            ):
                temp_glass_door = ''

            if ( row[7].value is not None
                    and '자동문' in row[7].value
            ):
                width_glassdoor = int(row[10].value * 1000)
                height_glassdoor = int(row[11].value * 1000)

                if ( row[13].value is None
                ):
                    standard_glassdoor = '편개'
                    temp_glass_door = f"{row[7].value} | {standard_glassdoor},{width_glassdoor}*{height_glassdoor} | SET"
                else:
                    standard_glassdoor = '양개'
                    temp_glass_door = f"{row[7].value} | {standard_glassdoor},{width_glassdoor}*{height_glassdoor} | SET"
            else:
                temp_glass_door = ''

            if ( row[7].value is not None
                    and '강화' in row[7].value
            ):
                width_glassdoor = int(row[10].value * 1000)
                height_glassdoor = int(row[11].value * 1000)
                temp_glass_door = f"{row[7].value}유리도어 | {row[8].value},{row[9].value}mm,{width_glassdoor}*{height_glassdoor},손보호용 | EA"

            if ( row[7].value is not None
                    and'로이삼중' in row[7].value
            ):
                width_glassdoor = int(row[10].value * 1000)
                height_glassdoor = int(row[11].value * 1000)

                if ( row[12].value is None
                ):
                    argon_glassdoor = ''
                    temp_glass_door = f"{row[7].value}유리도어{argon_glassdoor} | {row[8].value},{row[9].value}mm,{width_glassdoor}*{height_glassdoor} | EA"
                else:
                    argon_glassdoor = '(아르곤)'
                    temp_glass_door = f"{row[7].value}유리도어{argon_glassdoor} | {row[8].value},{row[9].value}mm,{width_glassdoor}*{height_glassdoor} | EA"
            else:
                temp_glass_door = ''
        except:
            print('유리문오류')


        # 방충망 insect_screen
        try:
            if ( row[16].value is not None
            ):
                temp_insect_screen = 'AL롤방충망 | 후레임포함 | M2'
            else:
                temp_insect_screen = ''
        except:
            print('방충망오류')


        item = ItemStandard(
            windows_name = temp_windows_name,
            glass_standard = temp_glass_standard,
            fire_entrance = temp_fire_entrance,
            glass_door = temp_glass_door,
            insect_screen = temp_insect_screen,
            remark = row[18].value,
            )
        items.append(item)

        # 시스템도어일때 도어 강제 삭제(오류방지용)
        if ( '시스템' in item.windows_name
        ):
            item.glass_door = ''



    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '창호완성'
    head_title = ['창호', '유리', '소방진입창', '도어', '방충망', 'Remark']
    new_sheet.append(head_title)
    new_sheet.freeze_panes = "A2"
    new_sheet.column_dimensions["A"].width = 50
    new_sheet.column_dimensions["B"].width = 50
    new_sheet.column_dimensions["C"].width = 50
    new_sheet.column_dimensions["D"].width = 50
    new_sheet.column_dimensions["E"].width = 30
    new_sheet.column_dimensions["F"].width = 12

    for item in items:
        new_sheet.append(item.to_excel())


    new_workbook.save("C:\\Users\ckddn\Desktop\창호완성.xlsx")



# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')


