from openpyxl import load_workbook, Workbook

from Architect.RC.ExcelStandard import ExcelStandard

from datetime import datetime

import platform
import subprocess

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')

# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0199_etc'
##################################

openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/토목.xlsx'
saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/토목-' + fileCreateDate + '.xlsx'


# openFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\창호.xlsx'
# saveFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\창호완성-' + fileCreateDate + '.xlsx'

# openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\창호.xlsx'
# saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\창호완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)

    내역목록 = []

    sheetname = "내역서BOQ"
    if sheetname in excel.sheetnames:
        worksheet = excel[sheetname]

        층확정 = ''
        호확정 = ''
        실확정 = ''
        품명확정 = ''
        규격확정 = ''
        부위확정 = ''
        산식확정 = ''
        수량확정 = ''

        for row in worksheet.iter_rows(min_row=4):

            공종값 = row[1].value
            규격값 = row[2].value
            단위값 = row[3].value
            수량값 = row[4].value
            비고값 = row[10].value

            if 비고값 is not None:

                비고값 = 비고값.replace('\n', '').replace('\r', '')
                공종값 = f'{공종값}/{비고값}'
                # 공종값 = 공종값 + '/' + 비고값

            if 공종값 is not None and 단위값 is not None:

                품명확정 = 공종값
                규격확정 = 규격값
                단위확정 = 단위값
                수량확정 = 수량값

                내역 = ExcelStandard(
                    층='',
                    호='',
                    실='',
                    대공종='토목',
                    중공종='',
                    코드='',
                    품명=품명확정,
                    규격=규격확정,
                    단위=단위확정,
                    부위='',
                    타입='외부',
                    산식=수량확정,
                    수량=수량확정,
                    Remark=''
                )
                내역목록.append(내역)

    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '토목완성'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 15
    new_sheet.column_dimensions["H"].width = 15

    for 내역 in 내역목록:
        new_sheet.append(내역.to_excel())

    new_workbook.save(saveFilePath)


    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs !='windows':
        subprocess.call(['open', saveFilePath])


if __name__ == '__main__':
    excel_normalize('PyCharm')
