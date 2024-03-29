

from openpyxl import load_workbook, Workbook


from Architect.Civil.ItemStandard import ItemStandard


def excel_normalize(name):
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\토목.xlsx',
        data_only=True)

    worksheet = excel['수량집계표']
    items = []
    name = ""
    for row in worksheet.iter_rows(min_col=0, max_col=29, min_row=4):
        # 단위없음 삭제
        if (row[15].value is None):
            continue

        # 품명
        if (row[0].value is not None):
            name = row[0].value



        item = ItemStandard(
            name = name.replace('\n',''),
            standard = row[9].value,
            unit = row[15].value,
            formula = row[17].value,
            sum = row[17].value,
            )
        items.append(item)




    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '토목(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    new_sheet.append(head_title)
    for item in items:
        new_sheet.append(item.to_excel())


    new_workbook.save("C:\\Users\ckddn\Desktop\토목완성.xlsx")

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
