import re

from openpyxl import load_workbook, Workbook

from Architect.TEST.QuantityItemStandard import QuantityItemStandard



def excel_normalize(name):
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\기계수량산출서.xlsx',
        data_only=True)

    worksheet = excel['목록별산출서']
    items = []
    temp_category = ""
    aaa = ""
    for row in worksheet.iter_rows(min_col=0, max_col=13, min_row=2):
        if (row[3].value is not None
            and row[1].value == '공종명Line'
        ):
            temp_category = row[3].value
            # print(temp_category)
            if '::' in temp_category:
                temp_levels = temp_category.split('::')[-1]
                if 'B' in temp_levels:
                    aaa = temp_levels.split(' ')[0] + 'F'
                elif 'F' in temp_levels:
                    aaa = temp_levels.split(' ')[0].replace('F','') + 'F'
                else:
                    aaa= ''

        if (row[4].value is not None
                and row[10].value is not None
        ):
            temp_formula = row[4].value

        if (row[7].value is None
            and row[8].value is None
            and row[9].value is None
        ):
            continue

        if (row[7].value == '명칭'
            and row[8].value == '규격'
            and row[9].value == '단위'
        ):
            continue

        item = QuantityItemStandard(
            floor = aaa,
            name = row[7].value,
            category = temp_category,
            standard = row[8].value,
            unit = row[9].value,
            formula = temp_formula,
            unit_formula = row[10].value,
            sum = row[12].value
            )
        items.append(item)

    for item in items:
        if '장비설치공사' in item.category:
            item.category = '장비설치공사'

        if '옥외배관공사' in item.category:
            item.category = '옥외배관공사'

        if '위생기구설치공사' in item.category:
            item.category = '위생기구설치공사'

        if '급수급탕배관공사' in item.category:
            item.category = '급수급탕배관공사'

        if '오배수배관공사' in item.category:
            item.category = '오배수배관공사'

        if '환기설비공사' in item.category:
            item.category = '환기덕트설치공사'



    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '기계(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    for item in items:
        new_sheet.append(item.to_excel())


    new_workbook.save("C:\\Users\ckddn\Desktop\기계완성.xlsx")

if __name__ == '__main__':
    excel_normalize('PyCharm')