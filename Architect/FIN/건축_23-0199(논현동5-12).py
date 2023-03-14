import os

import re

from openpyxl import load_workbook, Workbook

from Architect.FIN.ItemStandard import ItemStandard

from datetime import datetime

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')





# B1F_기계실, FSD_02    [문]   0.95*2.1       (  1EA)
# floor_regex = r"([PH]*[\d~]*[bB]*\d+F)_(.*)"
# B301 근린생활시설
# 101 근린생활시설

floor_regex = r"([A-Z]*[0-9]+)\s+(.*)"
# ASD_01 / 2000*2140 방화유리자동문포함(면적:4.28)
# window_regex = r"([A-Z_0-9]+)+\s/\s(\d+)\*(\d+).*\(면적:(\d+.\d+)\)"
# SD3       [문]   0.6*1.8        (  1EA)
window_regex = r"([A-Z]+\d+[A-Z]*)\s*\[[가-힣]\]\s*(\d+[.\d+]*)\*(\d+.\d+)\s*\(\s*([0-9])+.*"
# \s(\d+.\d+)*(\d+.\d+)\s\((\s\d+EA)\)



# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0031_etc'
##################################

openFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/건축.xlsx'
saveFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/건축완성-' + fileCreateDate + '.xlsx'

#openFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\건축.xlsx'
#saveFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\건축완성-' + fileCreateDate + '.xlsx'

#openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축.xlsx'
#saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축완성-' + fileCreateDate + '.xlsx'



def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath,data_only=True)

    items = []
    # index init
    position_index = 0
    name_index = 1
    standard_index = 2
    unit_index = 3
    formular_index = 5
    each_quantity_index = 6
    count_index = 7
    quantity_index = 8

    # 각종 변수들 초기화
    floor = location = room_name = type = ''
    sheet_name = '1. 본동-물량산출서'

    if sheet_name in excel.sheetnames:
        worksheet = excel[sheet_name]

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=2):

            if (row[position_index].value is not None
                    and (row[name_index].value is None or row[name_index].value == '')
                    and (row[standard_index].value is None or row[standard_index].value == '')
                    and (row[unit_index].value is None or row[unit_index].value == '')
                    and (row[formular_index].value is None or row[formular_index].value == '')
                    and (row[each_quantity_index].value is None or row[each_quantity_index].value == '')
                    and (row[count_index].value is None or row[count_index].value == '')
                    and (row[quantity_index].value is None or row[quantity_index].value == '')
            ):

                # ASD_01 / 2000*2140 방화유리자동문포함(면적:4.28)
                if is_window(row[position_index].value):
                    # name, width, height, area = window_name_stand(row[position_index].value)
                    name, width, height, count = window_name_stand(row[position_index].value)
                    item = ItemStandard(
                        floor=floor,
                        location=location,
                        roomname=room_name,
                        name=name,
                        # standard=f'{format(int(width) / 1000.0, ".3f")}*{format(int(height) / 1000.0, ".3f")}={format(float(area), ".3f")}',
                        standard=f'{format(float(width), ".3f")}*{format(float(height), ".3f")}={format(float(width)*float(height), ".3f")}',
                        unit="EA",
                        type=type,
                        formula=count,
                        sum=count,
                    )
                    items.append(item)
                    continue

                # 공사명 : 영등포동2가 다세대주택 및 오피스텔 신축공사 > 본관동 > 창호
                if ('>' in row[position_index].value and row[position_index].value != ''):
                    value_split = row[position_index].value.split('>')
                    type = value_split[-1].strip()
                    continue

                # B1F_기계실, FSD_02    [문]   0.95*2.1       (  1EA)
                if is_floor(row[position_index].value):
                    try:
                        split = row[position_index].value.split(' ')
                        if len(split) == 2:
                            floor = split[0]
                            location = split[1]
                        else:
                            location = split[0]
                        room_name = location
                    except:
                        print(f'{floor=}, {location=}')
                    continue

                # 정면도
                # [가로]X    0.95  [세로]Y     2.1  [면적]A   1.995 ==> pass
                if ('[' not in row[position_index].value and row[position_index].value != ''):
                    try:
                        floor = ''
                        location = row[position_index].value
                        room_name = row[position_index].value
                    except:
                        print(f'{value_split=}')
                    continue

            # 품명없음 삭제
            if (row[name_index].value is None
                    or row[name_index].value == ""
            ):
                continue

            # 위치	품명	규격	단위	부위	산출식	물량	개소	수량	비고
            if (row[position_index].value == '위치'
                    and row[name_index].value == '품명'
                    and row[standard_index].value == '규격'
                    and row[unit_index].value == '단위'
            ):
                continue

            if row[count_index].value is not None and int(str(row[count_index].value)) > 1:
                formula = f'{row[formular_index].value} * {row[count_index].value}'
            else:
                formula = row[formular_index].value

            item = ItemStandard(
                floor=floor,
                location=location,
                roomname=room_name,
                name=row[name_index].value,
                standard=row[standard_index].value,
                unit=row[unit_index].value,
                type=type,
                formula=formula,
                sum=row[quantity_index].value,
            )
            items.append(item)

    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30

    for item in items:
        new_sheet.append(item.to_excel())



    new_workbook.save(saveFilePath)


def floor_name(text: str):
    match = re.match(floor_regex, text)
    return match.groups()


def is_floor(text: str) -> bool:
    stripText = text.strip()
    if stripText == "":
        return False

    split = text.split(' ')
    # print(len(split))
    # print(text)
    if len(split) == 2 or len(split) == 1:
        return True
    else:
        return False


# def is_floor(text: str) -> bool:
    # match = re.match(floor_regex, text)
    # if match is None:
    #     return False
    # else:
    #     return match.lastindex == 2


def window_name_stand(text: str):
    match = re.match(window_regex, text)
    return match.groups()


def is_window(text: str) -> bool:
    match = re.match(window_regex, text)
    if match is None:
        return False
    else:
        return match.lastindex == 4


if __name__ == '__main__':
    excel_normalize('PyCharm')
