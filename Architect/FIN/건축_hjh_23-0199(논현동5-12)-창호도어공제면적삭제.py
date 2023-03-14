from openpyxl import load_workbook, Workbook

from Architect.FIN.ExcelStandard import ExcelStandard

from datetime import datetime

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')

# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0199_etc'
##################################

openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/건축.xlsx'
saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/건축완성-' + fileCreateDate + '.xlsx'


# openFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\창호.xlsx'
# saveFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\창호완성-' + fileCreateDate + '.xlsx'

# openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\창호.xlsx'
# saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\창호완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)

    내역목록 = []

    if '1. 본동-물량산출서' in excel.sheetnames:
        worksheet = excel['1. 본동-물량산출서']

        층확정 = ''
        호확정 = ''
        실확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        부위확정 = ''
        타입확정 = ''
        산식확정 = ''
        수량확정 = ''

        for row in worksheet.iter_rows(min_row=2):

            위치 = row[0].value
            품명 = row[1].value
            규격 = row[2].value
            단위 = row[3].value
            부위 = row[4].value
            산식 = row[5].value
            수량 = row[8].value

####### 타입 / 층 / 실명 뽑아내기  위치에만 데이터가 있는 항목들 대상  S ##############################

            if (위치 is not None
                    and (품명 is None or 품명 == '')
                    and (규격 is None or 규격 == '')
                    and (단위 is None or 단위 == '')
                    and (부위 is None or 부위 == '')
                    and (산식 is None or 산식 == '')
                    and (수량 is None or 수량 == '')
            ):

                위치분리빈칸 = 위치.split(' ')
                위치분리빈칸길이 = len(위치분리빈칸)

                위치분리슬래쉬 = 위치.split('/')
                위치분리슬래쉬길이 = len(위치분리슬래쉬)


                #print(위치, '/', 위치분리빈칸[0], '/', 위치분리빈칸길이)

                if '>' in 위치:
                    타입확정 = 위치.split('>')[-1]
                    continue

                if "/" in 위치:
                    층확정 = 위치분리슬래쉬[0].strip()
                    호확정 = 위치분리슬래쉬[0].strip()
                    실확정 = 위치분리슬래쉬[0].strip()

                    print(위치, '/', 위치분리슬래쉬[0], '/', 위치분리슬래쉬길이)
                    continue

                if '[문]' in 위치 or '[창]' in 위치:
                    continue

                # B301 근린생활   -> 층과 실로 분리하여 내역에 적용
                if 위치분리빈칸길이 == 1:
                    층확정 = ''
                    호확정 = 위치분리빈칸[0]
                    실확정 = 위치분리빈칸[0]
                    continue

                elif 위치분리빈칸길이 == 2:
                    층확정 = 위치분리빈칸[0]
                    호확정 = 위치분리빈칸[1]
                    실확정 = 위치분리빈칸[1]

                   # print(층확정, 호확정, 실확정)
                    continue
                elif ('[가로]X' in 위치 or '[둘레]L' in 위치):
                    continue
                else:

                    품명 = 'XXXXX'
                    규격 = '*****'
                    수량 = 777

                    #if 위치분리빈칸길이 > 1:

                        #for i, col in enumerate(위치분리빈칸):
                          # print(i, '/', col)

                        # print(층확정, 호확정, 실확정, 품명확정, 규격확정, 단위확정, 타입확정, 산식확정, 수량확정)

####### 타입 / 층 / 실명 뽑아내기  위치에만 데이터가 있는 항목들 대상  E ##############################


            if (위치 == '위치'
                    and 품명 == '품명'
                    and 규격 == '규격'
            ):
                continue


            if ((위치 is None or 위치 == '')
                    and (품명 is None or 품명 == '')
                    and (규격 is None or 규격 == '')
            ):
                continue

            if (('[가로]X' in 위치 or '[둘레]L' in 위치 or '공사명' in 위치)
                    and (품명 is None or 품명 == '')
                    and (규격 is None or 규격 == '')):
                continue

            if (품명 is not None and 수량 is not None):

                품명확정 = 품명
                규격확정 = 규격
                단위확정 = 단위
                부위확정 = row[4].value
                산식확정 = row[5].value
                수량확정 = row[8].value

                if 수량 is not None:
                    수량확정 = 수량

                내역 = ExcelStandard(
                    층=층확정,
                    호=호확정,
                    실=실확정,
                    대공종='건축',
                    중공종='',
                    코드='',
                    품명=품명확정,
                    규격=규격확정,
                    단위=단위확정,
                    부위=부위확정,
                    타입=타입확정,
                    산식=산식확정,
                    수량=수량확정,
                    Remark=''
                )
                내역목록.append(내역)

    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축완성'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 15
    new_sheet.column_dimensions["H"].width = 15

    for 내역 in 내역목록:
        new_sheet.append(내역.to_excel())
        new_workbook.save(saveFilePath)


if __name__ == '__main__':
    excel_normalize('PyCharm')
