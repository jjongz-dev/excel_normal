from openpyxl import load_workbook, Workbook

from Architect.FIN.ItemStandard import ItemStandard

from Architect.FIN.ItemStandard2 import ItemStandard2


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\건축.xlsx',
        data_only=True)

    items = []
    windows_dict = {}

    # temporary_constrution
    if '가설산출서' in excel.sheetnames:
        worksheet = excel['가설산출서']
        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=5):
            if (row[0].value is not None
                    and row[2].value is None
                    and row[3].value is None
                    and row[4].value is None
                    and row[5].value is None
                    and row[6].value is None
            ):
                try:
                    if '개소' in row[0].value:
                        temp_split_roomname = row[0].value.split('개소')[0].replace(' ','')
                        if '구분명' in temp_split_roomname:
                            temp_roomname = temp_split_roomname.split(':')[-1].strip(' ')
                        continue
                except Exception as e:
                    print('예외가 발생했습니다.', e)
                    print('가설산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[7].value is None
                    or row[7].value == "'"
                    or row[7].value == '0'
                    or row[7].value == 0
            ):
                continue

            item = ItemStandard(
                floor='1F',
                location=temp_roomname,
                roomname=temp_roomname,
                name=row[2].value,
                standard=row[3].value,
                unit=row[4].value,
                type='외부',
                formula=row[5].value,
                sum=row[7].value,
            )
            items.append(item)


    # earthwork
    if '토공산출서' in excel.sheetnames:
        worksheet = excel['토공산출서']
        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=4):
            if (row[0].value is not None
                    and row[3].value is None
                    and row[4].value is None
                    and row[5].value is None
                    and row[6].value is None
                    and row[8].value is None
            ):
                try:
                    temp_split_roomname = row[0].value.split('개소')[0]
                    if '구분명' in temp_split_roomname:
                        temp_roomname = temp_split_roomname.split(':')[-1].strip('[] ')
                    continue
                except Exception as e:
                    print('예외가 발생했습니다.', e)
                    print('토공산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[8].value is None
                    or row[8].value == "'"
                    or row[8].value == '0'
                    or row[8].value == 0
            ):
                continue

            item = ItemStandard(
                floor='1F',
                location=temp_roomname,
                roomname=temp_roomname,
                name=row[3].value,
                standard=row[4].value,
                unit=row[5].value,
                type='외부',
                formula=row[6].value,
                sum=row[8].value,
            )
            items.append(item)

    # internal_construction
    if '내부산출서' in excel.sheetnames:
        worksheet = excel['내부산출서']
        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=3):
            if (row[0].value is not None
                    and row[2].value is None
                    and row[3].value is None
                    and row[4].value is None
                    and row[5].value is None
                    and row[6].value is None
            ):
                try:
                    if '층' in row[0].value:
                        temp_levels = row[0].value.strip('[]').replace(' ', '')
                        temp_level = temp_levels.split('.')[-1]
                    if '실명' in row[0].value:
                        temp_roomnames = row[0].value.replace(' ', '')
                        temp_roomname = temp_roomnames.split('개소')[0].split('.')[-1]
                    continue
                except Exception as e:
                    print('예외가 발생했습니다.', e)
                    print('내부산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[6].value is None
                    or row[6].value == "'"
                    or row[6].value == '0'
                    or row[6].value == 0
                    or row[6].value == '물량'
            ):
                continue

            item = ItemStandard(
                floor=temp_level,
                location=temp_roomname,
                roomname=temp_roomname,
                name=row[2].value,
                standard=row[3].value,
                unit=row[4].value,
                type='내부',
                formula=row[5].value,
                sum=row[6].value,
            )
            items.append(item)

    # external_construction
    if '외부산출서' in excel.sheetnames:
        worksheet = excel['외부산출서']

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=5):
            if (row[0].value is not None
                and row[3].value is None
                and row[4].value is None
                and row[5].value is None
                and row[6].value is None
                and row[8].value is None
            ):
                try:
                    if '구분명' in row[0].value:
                        temp_roomnames = row[0].value.replace(' ', '')
                        temp_roomname = temp_roomnames.split('개소')[0].split('.')[-1]
                        continue
                except Exception as e:
                    print('예외가 발생했습니다', e)
                    print('외부산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[8].value is None
                    or row[8].value == "'"
                    or row[8].value == '0'
                    or row[8].value == 0
                    or row[8].value == '물량'
            ):
                continue

            item = ItemStandard(
                floor='',
                location=temp_roomname,
                roomname=temp_roomname,
                name=row[3].value,
                standard=row[4].value,
                unit=row[5].value,
                type='외부',
                formula=row[6].value,
                sum=row[8].value,
            )
            items.append(item)

    if '계단산출서' in excel.sheetnames:
        worksheet = excel['계단산출서']
        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=4):
            # 품명없음 삭제
            if (row[6].value is None
                    or row[6].value == "'"
                    or row[6].value == '0'
                    or row[6].value == 0
                    or row[6].value == '물량'
            ):
                continue

            item = ItemStandard(
                floor='',
                location='계단실',
                roomname='계단실',
                name=row[2].value,
                standard=row[3].value,
                unit=row[4].value,
                type='내부',
                formula=row[5].value,
                sum=row[6].value,
            )
            items.append(item)

    if '공용산출서' in excel.sheetnames:
        worksheet = excel['공용산출서']
        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=5):
            # 품명없음 삭제
            if (row[6].value is None
                    or row[6].value == "'"
                    or row[6].value == '0'
                    or row[6].value == 0
                    or row[6].value == '물량'
            ):
                continue

            item = ItemStandard(
                floor='1F',
                location='기타공사',
                roomname='기타공사',
                name=row[2].value,
                standard=row[3].value,
                unit=row[4].value,
                type='내부',
                formula=row[5].value,
                sum=row[6].value,
            )
            items.append(item)

    # List of windows and doors by Dong
    if '동별창호리스트' in excel.sheetnames:
        worksheet = excel['동별창호리스트']

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=5):
            if row[0].value == '' or len(row[0].value) < 2:
                break

            if row[1].value is not None:
                windows_dict[row[1].value] = row[10].value
                windows_standard1 = f'{(row[2].value):0.3f}'
                windows_standard2 = f'{(row[3].value):0.3f}'
                windows_standard3 = f'{(row[4].value):0.3f}'
                windows_standard = f"{windows_standard1}*{windows_standard2}={windows_standard3}"
                if row[9].value is not None and len(row[9].value) > 2:
                    windows_name = f"{row[1].value}({row[1].value})"
                else:
                    windows_name = f"{row[1].value}"

            item = ItemStandard(
                floor='',
                location='',
                roomname='',
                name=windows_name,
                standard=windows_standard,
                unit='EA',
                type='창호',
                formula=row[10].value,
                sum=row[10].value,
            )
            items.append(item)

    # windows
    if '창호산출서' in excel.sheetnames:
        worksheet = excel['창호산출서']

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=4):
            if (row[0].value is not None
                and row[1].value is None
                and row[2].value is None
                and row[3].value is None
                and row[4].value is None
                and row[5].value is None
            ):
                try:
                    temp_window_name = row[0].value.split('(')[0]
                    if '창호' in temp_window_name:
                        window_name = temp_window_name.split(':')[-1].strip(' ')
                        continue
                except Exception as e:
                    print('예외가 발생했습니다', e)
                    print('창호산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[5].value is None
                    or row[5].value == "'"
                    or row[5].value == '0'
                    or row[5].value == 0
                    or row[5].value == '물량'
            ):
                continue

            item = ItemStandard(
                floor='',
                location=window_name,
                roomname=window_name,
                name=row[1].value,
                standard=row[2].value,
                unit=row[3].value,
                type='창호',
                formula=f"({row[4].value})*<수량>({windows_dict[window_name]})",
                sum=float(row[5].value)*(float(windows_dict[window_name])),
            )
            items.append(item)


    items2 = []
    if '동별집계표' in excel.sheetnames:
        worksheet2 = excel['동별집계표']
        temp_constructionwork = ""
        for row in worksheet2.iter_rows(min_col=1, max_col=5, min_row=4):
            #중공종
            if (row[1].value.__contains__('내  역  삭  제')
            ):
                continue

            if (row[1].value is not None
                   and row[4].value is None
            ):
                temp_constructionwork = row[1].value.replace(" ","")


            item2 = ItemStandard2(
                constructionWork = temp_constructionwork,
                name = row[1].value,
                standard = row[2].value,
                unit = row[3].value,
                sum = row[4].value,
                )
            items2.append(item2)

    # for item in items:
    #     print(item)

# 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    for item in items:
        new_sheet.append(item.to_excel())

    sheet = new_workbook.create_sheet(title='집계표')
    sheet.append(['중공종', '품명', '규격', '단위', '수량(할증전)'])
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30
    for item2 in items2:
        sheet.append(item2.to_excel())

    new_workbook.save("C:\\Users\ckddn\Desktop\건축완성.xlsx")



# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')