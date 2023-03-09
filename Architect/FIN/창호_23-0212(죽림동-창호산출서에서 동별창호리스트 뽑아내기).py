from openpyxl import load_workbook, Workbook

from Architect.FIN.ItemStandard import ItemStandard

from Architect.FIN.ItemStandard2 import ItemStandard2

from Architect.FIN.PasingRule import Floorlevel, Deleteitem, Basicchange

from datetime import datetime

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')



# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0212_ko'
##################################

openFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/창호.xlsx'
saveFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/창호완성 -' + fileCreateDate + '.xlsx'

#openFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\창호.xlsx'
#saveFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\창호완성-' + fileCreateDate + '.xlsx'

#openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\창호.xlsx'
#saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\창호완성-' + fileCreateDate + '.xlsx'




def excel_normalize(name, column_dimensions=None):

    excel = load_workbook(openFilePath)

    items = []

    if '창호산출서' in excel.sheetnames:
        worksheet = excel['창호산출서']

        for row in worksheet.iter_rows(min_row=4):

            부위 = row[0].value
            품명 = row[1].value
            규격 = row[2].value
            단위 = row[3].value
            산식 = row[4].value
            물량 = row[5].value

            if (부위 is not None
                    and 품명 is None
                    and 규격 is None
                    and 단위 is None
                    and 산식 is None
                    and 물량 is None
            ):
                if '건축공사' in 부위:

                    창호명 = 부위.split('(')[0]
                    창호명 = 창호명.split(':')[-1].strip()

                    사이즈 = 부위.split('Size:')[1].split('공제면적')[0]

                    # 사이즈>>    7 * 4 = 28
                    가로 = f'{float(사이즈.split("*")[0]):0.3f}'

                    #가로 = f'{float(가로):0.3f}'

                    세로 = 사이즈.split('*')[1].split('=')[0]

                    면적 = 사이즈.split('*')[1].split('=')[-1]

                    print(창호명,'/', 사이즈, '/', 가로, 세로, 면적)












# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')


