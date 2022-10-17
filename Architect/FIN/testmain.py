from openpyxl import load_workbook, Workbook

from Architect.FIN.ItemStandard import ItemStandard

from Architect.FIN.PasingRule import Floorlevel


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\건축.xlsx',
        data_only=True)

    items = []
    levels = []
    floorsupportlevels = []
    temporary_row_index_list = []
    earthwork_row_index_list = []
    internal_row_index_list = []
    external_row_index_list = []
    dong_windowslist_row_index_list = []
    windowslist_row_index_list = []
    windows_dict = {}
    start_titles = ['부위', '도형', '구분']
    part_index = ""
    name_index = ""
    standard_index = ""
    unit_index = ""
    formular_index = ""
    quantity_index = ""
    figure_index = ""
    width_index = ""
    height_index = ""
    area_index = ""
    note_index = ""

    # temporary_constrution
    if '가설산출서' in excel.sheetnames:
        worksheet = excel['가설산출서']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    title_row_index = row_index
                    temporary_row_index_list.append(title_row_index)
                    break
            row_index += 1

        for index, column in enumerate(worksheet[temporary_row_index_list[0]]):
            match column.value:
                case '부위':
                    part_index = index
                case '품명':
                    name_index = index
                case '규격':
                    standard_index = index
                case '단위':
                    unit_index = index
                case '산식':
                    formular_index = index
                case '물량':
                    quantity_index = index

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=(temporary_row_index_list[0]+1)):
            if (row[part_index].value is not None
                    and row[name_index].value is None
                    and row[standard_index].value is None
                    and row[unit_index].value is None
                    and row[formular_index].value is None
                    and row[quantity_index].value is None
            ):
                try:
                    temp_split_roomname = row[part_index].value.split('개소')[0]
                    if '구분명' in temp_split_roomname:
                        temp_roomname = temp_split_roomname.split(':')[-1].strip('[] ')
                    continue
                except Exception as e:
                    print('예외가 발생했습니다.', e)
                    print('가설산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[quantity_index].value is None
                    or row[quantity_index].value == "'"
                    or row[quantity_index].value == '0'
                    or row[quantity_index].value == 0
            ):
                continue

            item = ItemStandard(
                floor='1F',
                location=temp_roomname,
                roomname=temp_roomname,
                name=row[name_index].value,
                standard=row[standard_index].value,
                unit=row[unit_index].value,
                type='내부',
                formula=row[formular_index].value,
                sum=row[quantity_index].value,
            )
            items.append(item)

    # earthwork
    if '토공산출서' in excel.sheetnames:
        worksheet = excel['토공산출서']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    title_row_index = row_index
                    earthwork_row_index_list.append(title_row_index)
                    break
            row_index += 1

        for index, column in enumerate(worksheet[earthwork_row_index_list[0]]):
            match column.value:
                case '도형':
                    figure_index = index
                case '품명':
                    name_index = index
                case '규격':
                    standard_index = index
                case '단위':
                    unit_index = index
                case '산식':
                    formular_index = index
                case '물량':
                    quantity_index = index


        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=(earthwork_row_index_list[0]+1)):
            if (row[figure_index].value is not None
                    and row[name_index].value is None
                    and row[standard_index].value is None
                    and row[unit_index].value is None
                    and row[formular_index].value is None
                    and row[quantity_index].value is None
            ):
                try:
                    temp_split_roomname = row[figure_index].value.split('개소')[0]
                    if '구분명' in temp_split_roomname:
                        temp_roomname = temp_split_roomname.split(':')[-1].strip('[] ')
                    continue
                except Exception as e:
                    print('예외가 발생했습니다.', e)
                    print('토공산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[quantity_index].value is None
                    or row[quantity_index].value == "'"
                    or row[quantity_index].value == '0'
                    or row[quantity_index].value == 0
            ):
                continue

            item = ItemStandard(
                floor='1F',
                location=temp_roomname,
                roomname=temp_roomname,
                name=row[name_index].value,
                standard=row[standard_index].value,
                unit=row[unit_index].value,
                type='외부',
                formula=row[formular_index].value,
                sum=row[quantity_index].value,
            )
            items.append(item)

    # internal_construction
    if '내부산출서' in excel.sheetnames:
        worksheet = excel['내부산출서']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    internal_row_index_list.append(row_index)
                    break
            row_index += 1

        for index, column in enumerate(worksheet[internal_row_index_list[0]]):
            match column.value:
                case '도형':
                    figure_index = index
                case '품명':
                    name_index = index
                case '규격':
                    standard_index = index
                case '단위':
                    unit_index = index
                case '산식':
                    formular_index = index
                case '물량':
                    quantity_index = index

        temp_levels = ""
        temp_roomname = ""
        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=(internal_row_index_list[0] - 1)):
            if (row[figure_index].value is not None
                    and row[name_index].value is None
                    and row[standard_index].value is None
                    and row[unit_index].value is None
                    and row[formular_index].value is None
                    and row[quantity_index].value is None
            ):
                try:
                    temp_split_levels = row[figure_index].value.split(' ')[-2]
                    if '층' in temp_split_levels:
                        temp_levels = temp_split_levels
                    temp_split_roomname = row[figure_index].value.split('개소')[0]
                    if '실명' in temp_split_roomname:
                        temp_roomname = temp_split_roomname.split(':')[-1].strip('[] ')
                    continue
                except Exception as e:
                    print('예외가 발생했습니다.', e)
                    print('내부산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[quantity_index].value is None
                    or row[quantity_index].value == "'"
                    or row[quantity_index].value == '0'
                    or row[quantity_index].value == 0
                    or row[quantity_index].value == '물량'
            ):
                continue

            item = ItemStandard(
                floor=temp_levels,
                location=temp_roomname,
                roomname=temp_roomname,
                name=row[name_index].value,
                standard=row[standard_index].value,
                unit=row[unit_index].value,
                type='내부',
                formula=row[formular_index].value,
                sum=row[quantity_index].value,
            )
            items.append(item)

    # external_construction
    if '외부산출서' in excel.sheetnames:
        worksheet = excel['외부산출서']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    external_row_index_list.append(row_index)
                    break
            row_index += 1

        for index, column in enumerate(worksheet[external_row_index_list[0]]):
            match column.value:
                case '도형':
                    figure_index = index
                case '품명':
                    name_index = index
                case '규격':
                    standard_index = index
                case '단위':
                    unit_index = index
                case '산식':
                    formular_index = index
                case '물량':
                    quantity_index = index

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=(external_row_index_list[0]+1)):
            if (row[figure_index].value is not None
                and row[name_index].value is None
                and row[standard_index].value is None
                and row[unit_index].value is None
                and row[formular_index].value is None
                and row[quantity_index].value is None
            ):
                try:
                    temp_split_roomname = row[figure_index].value.split('개소')[0]
                    if '구분명' in temp_split_roomname:
                        temp_roomname = temp_split_roomname.split(':')[-1].strip('[] ')
                        continue
                except Exception as e:
                    print('예외가 발생했습니다', e)
                    print('외부산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[quantity_index].value is None
                    or row[quantity_index].value == "'"
                    or row[quantity_index].value == '0'
                    or row[quantity_index].value == 0
                    or row[quantity_index].value == '물량'
            ):
                continue

            item = ItemStandard(
                floor='',
                location=temp_roomname,
                roomname=temp_roomname,
                name=row[name_index].value,
                standard=row[standard_index].value,
                unit=row[unit_index].value,
                type='외부',
                formula=row[formular_index].value,
                sum=row[quantity_index].value,
            )
            items.append(item)

    # List of windows and doors by Dong
    if '동별창호리스트' in excel.sheetnames:
        worksheet = excel['동별창호리스트']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    dong_windowslist_row_index_list.append(row_index)
                    break
            row_index += 1

        for index, column in enumerate(worksheet[dong_windowslist_row_index_list[0]]):
            match column.value:
                case '창호명':
                    name_index = index
                case '가로':
                    width_index = index
                case '세로':
                    height_index = index
                case '면적':
                    area_index = index
                case '비고':
                    note_index = index
                case '합계':
                    quantity_index = index

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=(dong_windowslist_row_index_list[0])+1):
            if row[0].value == '':
                break

            if row[name_index].value is not None:
                windows_name = f"{row[name_index].value}({row[note_index].value})"
                windows_dict[row[name_index].value] = row[quantity_index].value
                windows_standard1 = f'{(row[width_index].value):0.3f}'
                windows_standard2 = f'{(row[height_index].value):0.3f}'
                windows_standard3 = f'{(row[area_index].value):0.3f}'
                windows_standard = f"{windows_standard1}*{windows_standard2}={windows_standard3}"

            item = ItemStandard(
                floor='',
                location='',
                roomname='',
                name=windows_name,
                standard=windows_standard,
                unit='EA',
                type='창호',
                formula=row[quantity_index].value,
                sum=row[quantity_index].value,
            )
            items.append(item)

    # windows
    if '창호산출서' in excel.sheetnames:
        worksheet = excel['창호산출서']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    windowslist_row_index_list.append(row_index)
                    break
            row_index += 1

        for index, column in enumerate(worksheet[windowslist_row_index_list[0]]):
            match column.value:
                case '부위':
                    part_index = index
                case '품명':
                    name_index = index
                case '규격':
                    standard_index = index
                case '단위':
                    unit_index = index
                case '산식':
                    formular_index = index
                case '물량':
                    quantity_index = index

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=(windowslist_row_index_list[0]+1)):
            if (row[part_index].value is not None
                and row[name_index].value is None
                and row[standard_index].value is None
                and row[unit_index].value is None
                and row[formular_index].value is None
                and row[quantity_index].value is None
            ):
                try:
                    temp_window_name = row[part_index].value.split('(')[0]
                    if '창호' in temp_window_name:
                        window_name = temp_window_name.split(':')[-1].strip(' ')
                        continue
                except Exception as e:
                    print('예외가 발생했습니다', e)
                    print('창호산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[quantity_index].value is None
                    or row[quantity_index].value == "'"
                    or row[quantity_index].value == '0'
                    or row[quantity_index].value == 0
                    or row[quantity_index].value == '물량'
            ):
                continue

            item = ItemStandard(
                floor='',
                location=window_name,
                roomname=window_name,
                name=row[name_index].value,
                standard=row[standard_index].value,
                unit=row[unit_index].value,
                type='창호',
                formula=f"({row[formular_index].value})*<수량>({windows_dict[window_name]})",
                sum=float(row[quantity_index].value)*(float(windows_dict[window_name])),
            )
            items.append(item)



    for item in items:
        Floorlevel.launch(item, levels, floorsupportlevels)

    for item in items:
        if len(floorsupportlevels) > 1:
            if max(levels) > max(floorsupportlevels):
                if item.formula.__contains__('RF') or item.formula.__contains__('지붕'):
                    item.floor = str(max(levels)) + 'F'

#
    # for item in items:
    #     print(item)

# 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    for item in items:
        new_sheet.append(item.to_excel())

    new_workbook.save("C:\\Users\ckddn\Desktop\건축완성.xlsx")



# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')


