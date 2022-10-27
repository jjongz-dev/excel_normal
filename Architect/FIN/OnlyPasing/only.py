from openpyxl import load_workbook, Workbook

from Architect.FIN.ItemStandard import ItemStandard


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\건축완성.xlsx',
        data_only=True)

    items = []

    if '건축(데이터변경X)' in excel.sheetnames:
        worksheet = excel['건축(데이터변경X)']

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=2):

            item = ItemStandard(
                floor=row[0].value,
                location=row[1].value,
                roomname=row[2].value,
                name=row[6].value,
                standard=row[7].value,
                unit=row[8].value,
                type=row[10].value,
                formula=row[11].value,
                sum=row[12].value,
            )
            items.append(item)


# 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    for item in items:
        new_sheet.append(item.to_excel())

    new_workbook.save("C:\\Users\ckddn\Desktop\건축완성+파씽.xlsx")



# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')


