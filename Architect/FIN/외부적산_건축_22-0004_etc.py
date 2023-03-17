from openpyxl import load_workbook, Workbook

from Architect.FIN.ExcelStandard import ExcelStandard

from datetime import datetime

import platform
import subprocess

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')

# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '22-0004_etc'
##################################

openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/건축.xlsx'
saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/건축완성-' + fileCreateDate + '.xlsx'

# openFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\건축.xlsx'
# saveFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\건축완성-' + fileCreateDate + '.xlsx'

# openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축.xlsx'
# saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)

    내역목록 = []

    if '1. 본동-물량산출서' in excel.sheetnames:
        worksheet = excel['1. 본동-물량산출서']

        층확정 = ''
        호확정 = ''
        실확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        부위확정 = ''
        타입확정 = ''
        산식확정 = ''
        수량확정 = ''

        for row in worksheet.iter_rows(min_row=2):

            위치값 = row[0].value
            품명값 = row[1].value
            규격값 = row[2].value
            단위값 = row[3].value
            부위값 = row[4].value
            산식값 = row[5].value
            개소값 = row[7].value
            수량값 = row[8].value

            if (위치값 is not None and 위치값 != '') and (품명값 is None or 품명값 == '') and (규격값 is None or 규격값 == ''):
                if '>' in 위치값:
                    타입확정 = 위치값.split('>')[-1].strip()
                    continue

                if 위치값 is not None and (품명값 == None or 품명값 ==''):
                    층확정 = ''
                    호확정 = 위치값
                    실확정 = 위치값
                    continue

            if 위치값 == '위치' and 품명값 == '품명' and 규격값 == '규격':
                continue

            if (위치값 is None or 위치값 == '') and (품명값 is None or 품명값 == '') and (규격값 is None or 규격값 == ''):
                continue

            if 품명값 is not None and 수량값 is not None:

                품명확정 = 품명값
                규격확정 = 규격값
                단위확정 = 단위값
                부위확정 = 부위값
                산식확정 = 산식값
                수량확정 = 수량값
                개소확정 = 개소값

                내역 = ExcelStandard(
                    층=층확정,
                    호=호확정,
                    실=실확정,
                    대공종='건축',
                    중공종='',
                    코드='',
                    품명=품명확정,
                    규격=규격확정,
                    단위=단위확정,
                    부위=부위확정,
                    타입=타입확정,
                    산식=산식확정,
                    수량=수량확정,
                    Remark='',
                    개소=개소확정
                )
                내역목록.append(내역)

    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축완성'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark', '개소(확인용)']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 15
    new_sheet.column_dimensions["H"].width = 15

    for 내역 in 내역목록:
       new_sheet.append(내역.to_excel())

    new_workbook.save(saveFilePath)

    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])


if __name__ == '__main__':
    excel_normalize('PyCharm')

