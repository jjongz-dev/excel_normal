from openpyxl import load_workbook, Workbook

from Architect.FIN.ItemStandard import ItemStandard

from Architect.FIN.ItemStandard2 import ItemStandard2

from Architect.FIN.PasingRule import Floorlevel, Deleteitem, Basicchange

from datetime import datetime

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')



# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0212_ko'
##################################

openFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/창호.xlsx'
saveFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/창호완성 -' + fileCreateDate + '.xlsx'

#openFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\창호.xlsx'
#saveFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\창호완성-' + fileCreateDate + '.xlsx'

#openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\창호.xlsx'
#saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\창호완성-' + fileCreateDate + '.xlsx'




def excel_normalize(name, column_dimensions=None):

    excel = load_workbook(openFilePath,data_only=True)

    items = []
    windows_dict = {}

    if '창호산출서' in excel.sheetnames:
        worksheet = excel['창호산출서']
        windows_name = ''
        for row in worksheet.iter_rows(min_col=0, max_col=6, min_row=3):
            # 비교정보 가져오기
            if (row[0].value is not None
                    and row[1].value is None
                    and row[2].value is None
                    and row[3].value is None
                    and row[4].value is None
                    and row[5].value is None
            ):
                temp_name = row[0].value.split('(')[0]
                if '창호' in temp_name:
                    windows_name = temp_name.split(':')[-1].strip()
                continue

            # 품명없음 삭제
            if (row[5].value is None
                    or row[5].value == "'"
                    or row[5].value == '0'
                    or row[5].value == 0
                    or row[5].value == '물량'
            ):
                continue

            # print(f"keys : {windows_dict.keys()}")
            # print(f"request key : {windows_name}")
            item = ItemStandard(
                floor='',
                location=windows_name,
                roomname=windows_name,
                name=row[1].value,
                standard=row[2].value,
                unit=row[3].value,
                type='창호',
                formula=f"({row[4].value})*<수량>({windows_dict[windows_name]})",
                sum=float(row[5].value)*(float(windows_dict[windows_name])),
            )
            items.append(item)

    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    for item in items:
        new_sheet.append(item.to_excel())
        new_workbook.save(saveFilePath)


def float_format(column):
    return str(f'{float(column.value):0.3f}')


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')


