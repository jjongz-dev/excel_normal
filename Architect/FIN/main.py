from openpyxl import load_workbook, Workbook

from Architect.FIN.ItemStandard import ItemStandard

from Architect.FIN.ItemStandard2 import ItemStandard2

from Architect.FIN.PasingRule import Floorlevel, Deleteitem, Basicchange


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\건축.xlsx',
        data_only=True)


    items = []
    levels = []
    floorsupportlevels = []
    if excel.sheetnames.__contains__('산출근거집계표'):
        worksheet = excel['산출근거집계표']

        for row in worksheet.iter_rows(min_col=1, max_col=13, min_row=5):
            # 산출식 없음 삭제
            if ( row[2].value == '합        계'):
                continue

            # 구조이기 삭제
            if ( row[7].value == '구조이기'):
                continue

            # 0값 삭제
            if (row[12].value == '0' or row[12].value == 0 ):
                continue

            item = ItemStandard(
                floor = row[7].value,
                location = '',
                roomname=row[8].value,
                name = row[2].value,
                standard= row[3].value,
                unit=row[4].value,
                type=row[5].value,
                formula = row[9].value,
                sum=row[12].value,
                )
            items.append(item)

        for item in items:
            Basicchange.launch(item)

            Deleteitem.launch(item)

            Floorlevel.launch(item, levels, floorsupportlevels)

        for item in items:
            if len(floorsupportlevels) > 1:
                if max(levels) > max(floorsupportlevels):
                    if item.formula.__contains__('RF') or item.formula.__contains__('지붕'):
                        item.floor = str(max(levels)) + 'F'

    else:
        if excel.sheetnames.__contains__('가설산출서'):
            worksheet = excel['가설산출서']
            for row in worksheet.iter_rows(min_col=0, max_col=8, min_row=5):
                # 품명없음 삭제
                if row[2].value is None or row[2].value == "'" or row[7].value == '0' or row[7].value == 0:
                    continue

                item = ItemStandard(
                    floor='1F',
                    location='공통가설',
                    roomname='공통가설',
                    name=row[2].value,
                    standard=row[3].value,
                    unit=row[4].value,
                    type='내부',
                    formula=row[5].value,
                    sum=row[6].value,
                )
                items.append(item)

        if excel.sheetnames.__contains__('토공산출서'):
            worksheet = excel['토공산출서']
            for row in worksheet.iter_rows(min_col=0, max_col=9, min_row=5):
                # 품명없음 삭제
                if row[3].value is None or row[3].value == "'" or row[8].value == '0' or row[8].value == 0:
                    continue


                item = ItemStandard(
                    floor='FT',
                    location='기초하부',
                    roomname='기초하부',
                    name=row[3].value,
                    standard=row[4].value,
                    unit=row[5].value,
                    type='외부',
                    formula=row[6].value,
                    sum=row[7].value,
                )
                items.append(item)

        inlevels = ""
        inroomname = ""
        if '내부산출서' in excel.sheetnames:
        # if excel.sheetnames.__contains__('내부산출서'):
            worksheet = excel['내부산출서']
            for row in worksheet.iter_rows(min_col=0, max_col=7, min_row=3):
                # 층정보가져오기
                if (row[0].value is not None
                        and row[1].value is None
                        and row[2].value is None
                        and row[3].value is None
                        and row[4].value is None
                        and row[5].value is None
                ):
                    temp_level = row[0].value.split(' ')[-2]
                    if '층' in temp_level:
                        inlevels = temp_level
                    temp_inroomname = row[0].value.split('개소')[0]
                    if '실명' in temp_inroomname:
                        inroomname = temp_inroomname.split(':')[-1]
                    continue

                # 0값 삭제
                if row[6].value == '0' or  row[6].value == 0:
                    continue

                if row[2].value == '품명':
                    continue


                item = ItemStandard(
                    floor=inlevels,
                    location='',
                    roomname=inroomname,
                    name=row[2].value,
                    standard=row[3].value,
                    unit=row[4].value,
                    type='내부',
                    formula=row[5].value,
                    sum=row[6].value,
                )
                items.append(item)





    items2 = []
    if excel.sheetnames.__contains__('동별집계표'):
        worksheet2 = excel['동별집계표']
        constructionWork = ""
        for row in worksheet2.iter_rows(min_col=1, max_col=5, min_row=4):
            #중공종
            if (row[1].value.__contains__('내  역  삭  제')
            ):
                continue

            if (row[1].value is not None
                   and row[4].value is None
            ):
                constructionWork = row[1].value.replace(" ","")


            item2 = ItemStandard2(
                constructionWork = constructionWork,
                name = row[1].value,
                standard = row[2].value,
                unit = row[3].value,
                sum = row[4].value,
                )
            items2.append(item2)


    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    for item in items:
        new_sheet.append(item.to_excel())

    sheet = new_workbook.create_sheet(title='집계표')
    sheet.append(['중공종', '품명', '규격', '단위', '수량(할증전)'])
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30
    for item2 in items2:
        sheet.append(item2.to_excel())


    new_workbook.save("C:\\Users\ckddn\Desktop\건축완성.xlsx")

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')


