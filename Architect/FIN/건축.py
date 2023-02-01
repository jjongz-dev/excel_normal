from openpyxl import load_workbook, Workbook
from Architect.FIN.ItemStandard import ItemStandard
from Architect.FIN.ItemStandard2 import ItemStandard2
from Architect.FIN.PasingRule import Floorlevel, Deleteitem, Basicchange
from datetime import datetime

fileNameDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M%S')


# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0046'
##################################


openFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/건축.xlsx'
saveFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/건축완성-' + fileNameDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath, data_only=True)

    items = []
    levels = []
    floorsupportlevels = []
    if '산출근거집계표' in excel.sheetnames:
        worksheet = excel['산출근거집계표']

        for row in worksheet.iter_rows(min_col=1, max_col=13, min_row=5):
            # 산출식 없음 삭제
            if ( row[2].value == '합        계'):
                continue

            # 구조이기 삭제
            if ( row[7].value == '구조이기'):
                continue

            # 0값 삭제
            if (row[12].value == '0' or row[12].value == 0 ):
                continue

            item = ItemStandard(
                floor = row[7].value,
                location = '',
                roomname=row[8].value,
                name = row[2].value,
                standard= row[3].value,
                unit=row[4].value,
                type=row[5].value,
                formula = row[9].value,
                sum=row[12].value,
                )
            items.append(item)

        for item in items:
            Basicchange.launch(item)

            Deleteitem.launch(item)

            Floorlevel.launch(item, levels, floorsupportlevels)



        for item in items:
            if len(floorsupportlevels) > 1:
                if max(levels) > max(floorsupportlevels):
                    if item.formula.__contains__('RF') or item.formula.__contains__('지붕'):
                        item.floor = str(max(levels)) + 'F'

    else:
        if '가설산출서' in excel.sheetnames:
            worksheet = excel['가설산출서']
            temproomname = ""
            for row in worksheet.iter_rows(min_col=0, max_col=8, min_row=5):
                # 층정보가져오기
                if (row[0].value is not None
                        and row[1].value is None
                        and row[2].value is None
                        and row[3].value is None
                        and row[4].value is None
                        and row[5].value is None
                        and row[6].value is None
                        and row[7].value is None
                ):
                    temp_roomname = row[0].value.split('개소')[0]
                    if '구분명' in temp_roomname:
                        temproomname = temp_roomname.split(':')[-1].replace("[","").replace("]","").replace(" ","")
                    continue

                # 품명없음 삭제
                if (row[7].value is None
                        or row[7].value == "'"
                        or row[7].value == '0'
                        or row[7].value == 0
                        or row[7].value == '물량'
                ):
                    continue
                print(temproomname)
                item = ItemStandard(
                    floor='1F',
                    location=temproomname,
                    roomname=temproomname,
                    name=row[2].value,
                    standard=row[3].value,
                    unit=row[4].value,
                    type='내부',
                    formula=row[5].value,
                    sum=row[7].value,
                )
                items.append(item)

        if '토공산출서' in excel.sheetnames:
            worksheet = excel['토공산출서']
            for row in worksheet.iter_rows(min_col=0, max_col=9, min_row=5):
                # 품명없음 삭제
                if (row[8].value is None
                        or row[8].value == "'"
                        or row[8].value == '0'
                        or row[8].value == 0
                        or row[8].value == '물량'
                ):
                    continue

                item = ItemStandard(
                    floor='FT',
                    location='기초하부',
                    roomname='기초하부',
                    name=row[3].value,
                    standard=row[4].value,
                    unit=row[5].value,
                    type='외부',
                    formula=row[6].value,
                    sum=row[8].value,
                )
                items.append(item)

        if '내부산출서' in excel.sheetnames:
            worksheet = excel['내부산출서']
            inlevels = ""
            inroomname = ""
            for row in worksheet.iter_rows(min_col=0, max_col=7, min_row=3):
                # 층정보가져오기
                if (row[0].value is not None
                        and row[1].value is None
                        and row[2].value is None
                        and row[3].value is None
                        and row[4].value is None
                        and row[5].value is None
                ):
                    temp_level = row[0].value.split(' ')[-2]
                    if '층' in temp_level:
                        inlevels = temp_level
                    temp_roomname = row[0].value.split('개소')[0]
                    if '실명' in temp_roomname:
                        inroomname = temp_roomname.split(':')[-1]
                    continue

                # 품명없음 삭제
                if (row[6].value is None
                        or row[6].value == "'"
                        or row[6].value == '0'
                        or row[6].value == 0
                        or row[6].value == '물량'
                ):
                    continue

                item = ItemStandard(
                    floor=inlevels,
                    location=inroomname,
                    roomname=inroomname,
                    name=row[2].value,
                    standard=row[3].value,
                    unit=row[4].value,
                    type='내부',
                    formula=row[5].value,
                    sum=row[6].value,
                )
                items.append(item)

        if '외부산출서' in excel.sheetnames:
            worksheet = excel['외부산출서']
            exroomname = ""
            for row in worksheet.iter_rows(min_col=0, max_col=9, min_row=3):
                # 층정보가져오기
                if (row[0].value is not None
                        and row[1].value is None
                        and row[2].value is None
                        and row[3].value is None
                        and row[4].value is None
                        and row[5].value is None
                        and row[6].value is None
                        and row[7].value is None
                ):
                    temp_roomname = row[0].value.split('개소')[0]
                    if '구분명' in temp_roomname:
                        exroomname = temp_roomname.split(':')[-1]
                    continue

                # 품명없음 삭제
                if (row[8].value is None
                        or row[8].value == "'"
                        or row[8].value == '0'
                        or row[8].value == 0
                        or row[8].value == '물량'
                ):
                    continue

                item = ItemStandard(
                    floor='',
                    location=exroomname,
                    roomname=exroomname,
                    name=row[3].value,
                    standard=row[4].value,
                    unit=row[5].value,
                    type='외부',
                    formula=row[6].value,
                    sum=row[8].value,
                )
                items.append(item)

        if '조적산출서' in excel.sheetnames:
            worksheet = excel['조적산출서']
            for row in worksheet.iter_rows(min_col=0, max_col=6, min_row=3):
                # 층정보가져오기
                if (row[0].value is not None
                        and row[1].value is None
                        and row[2].value is None
                        and row[3].value is None
                        and row[4].value is None
                        and row[5].value is None
                ):
                    temp_level = row[0].value.split(' ')[-2]
                    if '층' in temp_level:
                        bricklevels = temp_level
                    temp_roomname = row[0].value.split('개소')[0]
                    if '실명' in temp_roomname:
                        brickroomname = temp_roomname.split(':')[-1]
                    continue

                # 품명없음 삭제
                if (row[5].value is None
                        or row[5].value == "'"
                        or row[5].value == '0'
                        or row[5].value == 0
                        or row[5].value == '물량'
                ):
                    continue

                item = ItemStandard(
                    floor=bricklevels,
                    location=brickroomname,
                    roomname=brickroomname,
                    name=row[1].value,
                    standard=row[2].value,
                    unit=row[3].value,
                    type='내부',
                    formula=row[4].value,
                    sum=row[5].value,
                )
                items.append(item)

        windows_dict = {}
        if '동별창호리스트' in excel.sheetnames:
            worksheet = excel['동별창호리스트']
            for row in worksheet.iter_rows(min_col=0, max_col=11, min_row=5):
                if row[0].value == '':
                    break

                if row[1].value is not None:
                    windows_name = row[1].value + '(' + row[9].value + ')'
                    windows_dict[row[1].value] = row[10].value
                    d1 = float_format(row[2])
                    d2 = float_format(row[3])
                    d3 = float_format(row[4])
                    windows_standard = d1 + '*' + d2 + '=' + d3

                item = ItemStandard(
                    floor='',
                    location='',
                    roomname='',
                    name=windows_name,
                    standard=windows_standard,
                    unit='EA',
                    type='창호',
                    formula=row[10].value,
                    sum=row[10].value,
                )
                items.append(item)


        if '창호산출서' in excel.sheetnames:
            worksheet = excel['창호산출서']
            windows_name = ''
            for row in worksheet.iter_rows(min_col=0, max_col=6, min_row=3):
                # 비교정보 가져오기
                if (row[0].value is not None
                        and row[1].value is None
                        and row[2].value is None
                        and row[3].value is None
                        and row[4].value is None
                        and row[5].value is None
                ):
                    temp_name = row[0].value.split('(')[0]
                    if '창호' in temp_name:
                        windows_name = temp_name.split(':')[-1].strip()
                    continue

                # 품명없음 삭제
                if (row[5].value is None
                        or row[5].value == "'"
                        or row[5].value == '0'
                        or row[5].value == 0
                        or row[5].value == '물량'
                ):
                    continue

                # print(f"keys : {windows_dict.keys()}")
                # print(f"request key : {windows_name}")
                item = ItemStandard(
                    floor='',
                    location='',
                    roomname='',
                    name=row[1].value,
                    standard=row[2].value,
                    unit=row[3].value,
                    type='창호',
                    formula=f"({row[4].value})*<수량>({windows_dict[windows_name]})",
                    sum=float(row[5].value)*(float(windows_dict[windows_name])),
                )
                items.append(item)

        # for key, value in windows_dict.items():
        #     print(f"{key}:{value}")


    # 신규아이템 가설 추가
    for numbering in range(3):
        numbering = numbering + 1
        temp_names = '가설신규아이템만#'
        names = f"{temp_names}{str(numbering)}"

        item = ItemStandard(
            floor='1F',
            location='가설',
            roomname='가설',
            name=names,
            standard='',
            unit='EA',
            type='내부',
            formula=float(1),
            sum=float(1),
        )
        items.append(item)



    items2 = []
    if excel.sheetnames.__contains__('동별집계표'):
        worksheet2 = excel['동별집계표']
        temp_constructionwork = ""
        for row in worksheet2.iter_rows(min_col=1, max_col=5, min_row=4):
            #중공종
            if (row[1].value.__contains__('내  역  삭  제')
            ):
                continue

            if (row[1].value is not None
                   and row[4].value is None
            ):
                temp_constructionwork = row[1].value.replace(" ","")


            item2 = ItemStandard2(
                constructionWork = temp_constructionwork,
                name = row[1].value,
                standard = row[2].value,
                unit = row[3].value,
                sum = row[4].value,
                )
            items2.append(item2)


    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    for item in items:
        new_sheet.append(item.to_excel())

    sheet = new_workbook.create_sheet(title='집계표')
    sheet.append(['중공종', '품명', '규격', '단위', '수량(할증전)'])
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30
    for item2 in items2:
        sheet.append(item2.to_excel())


    new_workbook.save(saveFilePath)


def float_format(column):
    return str(f'{float(column.value):0.3f}')


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')


