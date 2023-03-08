import os

import win32com.client
from openpyxl import load_workbook, Workbook

from Architect.RC.ItemStandard import ItemStandard

desktop_location = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

items = []
def excel_normalize(name, column_dimensions=None):
    print(name)
    excel = load_workbook(
        desktop_location + '\\구조.xlsx',
        read_only=True,
        data_only=True)

    # index init
    floor_part_index = 0

    concrete_standard_index = 4
    concrete_formular_index = 5
    concrete_quantity_index = 10

    formwork_standard_index = 11
    formwork_formular_index = 13
    formwork_quantity_index = 18

    rebar_standard_index = 21
    rebar_formular_index = 23
    rebar_quantity_index = 30


    # 각종 변수들 초기화
    floor = location = room_name = part = ''
    type = '구조'
    sheet_name = '본관동-물량산출서(옹벽)'
    if sheet_name in excel.sheetnames:
        worksheet = excel[sheet_name]

        min_row = 2
        for index, row in enumerate(worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=min_row)):

            if is_empty_row(row):
                continue
            if row[4].value == '종류' and row[5].value == '산출식':
                continue
            if row[4].value == '콘크리트(M3)':
                continue
            if row[0].value in ['계', '[비  고]']:
                continue
            if row[17].value in ['D', 'H D' , 'H D/s', 'SHD/s', 'UHD', 'SSH']:
                continue


            # print(index)

            if (row[floor_part_index].value is not None
                    and (row[concrete_formular_index].value is None or row[concrete_formular_index].value == '')
                    and (row[formwork_formular_index].value is None or row[formwork_formular_index].value == '')
                    and (row[rebar_formular_index].value is None or row[rebar_formular_index].value == '')

            ):
                # [ 본관동   기초 ]
                floor_part_name = row[floor_part_index].value
                if floor_part_name.startswith("[") and floor_part_name.endswith("]"):
                    try:
                        location = floor_part_name.replace('[', '').replace(']', '').strip().split('   ')[1]
                    except:
                        print(floor_part_name)
                    continue

            # CONC
            if row[concrete_standard_index].value is not None and 'Kg' in row[concrete_standard_index].value and row[concrete_formular_index].value is not None:
                if row[floor_part_index].value is not None:
                    next_row = worksheet[index + min_row + 1]
                    prev_row = worksheet[index + min_row - 1]
                    if next_row[floor_part_index].value is not None:
                        floor = next_row[floor_part_index].value
                        part = row[floor_part_index].value
                        print(index, floor, part)
                    if prev_row[floor_part_index].value is not None:
                        floor = row[floor_part_index].value
                        part = prev_row[floor_part_index].value
                        print(index, floor, part)

                formular = row[concrete_formular_index].value
                quantity = row[concrete_quantity_index].value

                if row[concrete_quantity_index].value is None:
                    next_row = worksheet[index + min_row + 1]
                    formular += str(next_row[concrete_formular_index].value)
                    if next_row[concrete_quantity_index].value is None:
                        next_next_row = worksheet[index + min_row + 2]
                        formular += str(next_next_row[concrete_formular_index].value)
                        quantity = next_next_row[concrete_quantity_index].value
                    else:
                        quantity = next_row[concrete_quantity_index].value

                item = ItemStandard(
                    floor=floor,
                    location=location,
                    name='콘크리트',
                    standard=row[concrete_standard_index].value,
                    part=part,
                    formula=formular,
                    sum=quantity,
                )
                items.append(item)

            # 거푸집
            if row[formwork_standard_index].value is not None and row[formwork_formular_index].value is not None:
                formular = row[formwork_formular_index].value
                quantity = row[formwork_quantity_index].value
                if row[formwork_quantity_index].value is None:
                    next_row = worksheet[index + min_row + 1]
                    formular += str(next_row[formwork_formular_index].value)
                    if next_row[formwork_quantity_index].value is None:
                        next_next_row = worksheet[index + min_row + 2]
                        formular += str(next_next_row[formwork_formular_index].value)
                        quantity = next_next_row[formwork_quantity_index].value
                    else:
                        quantity = next_row[formwork_quantity_index].value

                item = ItemStandard(
                    floor=floor,
                    location=location,
                    name='거푸집',
                    standard=row[formwork_standard_index].value,
                    part=part,
                    formula=formular,
                    sum=quantity,
                )
                items.append(item)

            print(f"맨앞={row[floor_part_index].value}", f"규격={row[rebar_standard_index].value}", f"산식={row[rebar_formular_index].value}")
            # 철근
            if row[rebar_standard_index].value is not None and 'HD' in row[rebar_standard_index].value and row[rebar_formular_index].value is not None:
                formular = row[rebar_formular_index].value
                quantity = row[rebar_quantity_index].value


                if row[rebar_quantity_index].value is None:
                    next_row = worksheet[index + min_row + 1]
                    formular += str(next_row[rebar_formular_index].value)
                    if next_row[rebar_quantity_index].value is None:
                        next_next_row = worksheet[index + min_row + 2]
                        formular += str(next_next_row[rebar_formular_index].value)
                        quantity = next_next_row[rebar_quantity_index].value
                    else:
                        quantity = next_row[rebar_quantity_index].value

                item = ItemStandard(
                    floor=floor,
                    location=location,
                    name='철근',
                    standard=row[rebar_standard_index].value,
                    part=part,
                    formula=formular,
                    sum=quantity,
                )
                items.append(item)

    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    try:
        for item in items:
            new_sheet.append(item.to_excel())
    except:
        print(item)

    excel = win32com.client.Dispatch('Excel.Application')
    try:
        result_excel_file_name = desktop_location + "\\구조완성.xlsx"
        new_workbook.save(result_excel_file_name)
        excel.Workbooks.Open(result_excel_file_name)
    except PermissionError as pe:
        for workbook in excel.Workbooks:
            full_name = workbook.FullName
            if full_name == result_excel_file_name:
                workbook.Close(True)
                new_workbook.save(result_excel_file_name)
                excel.Workbooks.Open(result_excel_file_name)


def is_empty_row(row):
    for column in row:
        if column.value is None:
            continue
        elif column.value is not None:
            return False
        elif column.value.strip() != '':
            return False
    return True

if __name__ == '__main__':
    excel_normalize('PyCharm')
