

from openpyxl import load_workbook, Workbook


from Architect.RC.ItemStandard import ItemStandard


def excel_normalize(name):
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\구조.xlsx',
        data_only=True)
    # names = excel.get_sheet_names()
    # print(names)
    worksheet = excel['부재별산출서']
    items = []
    floor = ""
    location = ""
    part = ""
    for row in worksheet.iter_rows(min_col=0, max_col=6, min_row=4):
        # 호
        if ( row[0].value is not None
                and row[1].value is None
                and row[2].value is None
                and row[3].value is None
                and row[4].value is None
                and row[5].value is None
        ):
            location = row[0].value.split('-')[-1].strip()
            continue

        # 층, 부위
        if ( row[0].value is not None
            and row[1].value is not None):
            floor = row[0].value
            part = row[1].value

        # 비고 제외
        if( row[2].value == '[ 비 고 ]'):
            continue

        # 콘크리트 규격 정규화 25-18-8 > 25-18-08
        concs = row[3].value.split('-')
        newconc = row[3].value
        if( len(concs) == 3):
            slope = concs[-1].zfill(2)
            newconc = '-'.join([concs[0], concs[1], slope])



        item = ItemStandard(
            floor = floor,
            location = location,
            name = row[2].value,
            standard= newconc,
            part = part,
            formula = row[4].value,
            sum = row[5].value,
            )
        # print(item.to_excel())
        items.append(item)


    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    new_sheet.append(head_title)
    for item in items:
        new_sheet.append(item.to_excel())


    new_workbook.save("C:\\Users\ckddn\Desktop\구조완성.xlsx")

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
