import re

from openpyxl import load_workbook, Workbook


from Architect.RC.UMID.ItemStandard import ItemStandard


def excel_normalize(name):
    excel = load_workbook(
        'C:\\Users\ckddn\Desktop\구조.xlsx',
        data_only=True)
    # names = excel.get_sheet_names()
    # print(names)
    items = []
    for sheetname in excel.sheetnames:
        worksheet = excel[sheetname]
        location = ""
        part = ""
        for row in worksheet.iter_rows(min_col=0, max_col=8, min_row=4):
            # 호
            if (row[0].value is not None
                    and row[1].value is not None):
                location = row[1].value.split(':')[-1]

            # 행삭제
            if (row[4].value is None):
                continue

            # 부위
            if (row[1].value is not None
                    and row[4].value is not None):
                part = row[1].value

            item = ItemStandard(
                floor=sheetname,
                location=location,
                name=row[2].value,
                standard=row[3].value,
                part=part,
                formula=row[5].value,
                sum=row[6].value,
            )
            # print(item.to_excel())
            items.append(item)

    for item in items:
        # 층정리
        if (re.match('\\d{1,2}', item.floor)):
            item.floor = re.sub(r'[^0-9]', '', item.floor).lstrip("0") + 'F'

        if (re.match('B\\d{1,2}', item.floor)):
            item.floor = 'B' + re.sub(r'[^0-9]', '', item.floor).lstrip("0") + 'F'

        if (re.match('P\\d{1,2}', item.floor)):
            item.floor = 'PH' + re.sub(r'[^0-9]', '', item.floor).lstrip("0") + 'F'

        if (item.floor == 'FTPIT'):
            item.floor = 'B2F'

    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    new_sheet.append(head_title)
    for item in items:
        new_sheet.append(item.to_excel())


    new_workbook.save("C:\\Users\ckddn\Desktop\구조완성.xlsx")

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
