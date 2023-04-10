from openpyxl import load_workbook, Workbook
from Architect.RC.ItemStandard import ItemStandard
from datetime import datetime

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')


# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '22-1178'
##################################

openFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/구조.xlsx'
saveFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/구조완성-' + fileCreateDate + '.xlsx'

#openFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\구조.xlsx'
#saveFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\구조완성-' + fileCreateDate + '.xlsx'

#openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조.xlsx'
#saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조완성-' + fileCreateDate + '.xlsx'



def excel_normalize(name):
    excel = load_workbook(openFilePath, data_only=True)

    # names = excel.get_sheet_names()
    # print(names)

    # 시트 추가가 필요한경우 오른쪽과 같이 추가 :   ,'시트명'
    # 추가해놓은 시트가 없는경우 자동으로 다음시트로 넘어감.
    sheet_names = ['부재별산출서', '아파트옹벽 Unit별산출서']
    items = []

    for sheet in sheet_names:
        if sheet in excel.sheetnames:
            worksheet = excel[sheet]
            location = ""
            floor = ""
            part = ""
            crossname = ""
            for row in worksheet.iter_rows(min_col=0, max_col=6, min_row=4):
                # 호
                if ( row[0].value is not None
                        and row[1].value is None
                        and row[2].value is None
                        and row[3].value is None
                        and row[4].value is None
                        and row[5].value is None
                ):
                    location = row[0].value.split('-')[-1].strip()
                    continue

                # 층, 부위, 이름
                if ( row[0].value is not None
                    and row[1].value is not None):
                    part = row[1].value
                    if row[0].value == 'FT':
                        floor = row[0].value
                    else :
                        floor = row[0].value + 'F'

                # 컷근이름 가져오기
                if ( row[2].value is not None
                        and row[5].value is not None):
                    temp_crossname = row[2].value.replace("'","")
                    if len(temp_crossname) >= 2:
                        crossname = temp_crossname


                # 비고 제외
                if( row[2].value == '[ 비 고 ]'):
                    continue

                # 콘크리트 규격 정규화 25-18-8 > 25-18-08
                concs = row[3].value.split('-')
                newconc = row[3].value
                if( len(concs) == 3):
                    slope = concs[-1].zfill(2)
                    newconc = '-'.join([concs[0], concs[1], slope])



                item = ItemStandard(
                    floor = floor,
                    location = location,
                    name = crossname,
                    standard= newconc,
                    part = part,
                    formula = row[4].value,
                    sum = row[5].value,
                    )
                # print(item.to_excel())
                items.append(item)



    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    new_sheet.append(head_title)
    for item in items:
        new_sheet.append(item.to_excel())


    new_workbook.save(saveFilePath)

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
