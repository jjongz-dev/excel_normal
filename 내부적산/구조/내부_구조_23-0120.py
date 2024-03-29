from openpyxl import load_workbook, Workbook
from ExcelStandard import ExcelStandard
from datetime import datetime
import platform
import subprocess

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')
systemOs = platform.system()


# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0120'
##################################


if systemOs == 'Darwin':
    openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/구조.xlsx'
    saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/구조완성-' + fileCreateDate + '.xlsx'
else:
    openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조.xlsx'
    saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)

    산출서목록 = []

    파싱시트목록 = ['부재별산출서', '기타산출서', '아파트옹벽 Unit별산출서']
    print('=================================')
    for 시트명 in 파싱시트목록:
        if 시트명 in excel.sheetnames:
            print("파싱시트체크 : ", 시트명, '(O)')
        else:
            print("파싱시트체크 : ", 시트명, '(X) - 확인필요')
    print('=================================')

    for sheetname in 파싱시트목록:

        if sheetname in excel.sheetnames:
            worksheet = excel[sheetname]

            층확정 = ''
            호확정 = ''
            실확정 = ''
            품명확정 = ''
            규격확정 = ''
            부위확정 = ''
            산식확정 = ''
            수량확정 = ''

            for row in worksheet.iter_rows(min_row=4):

                층값 = str(row[0].value)
                부호값 = row[1].value
                명칭값 = row[2].value
                규격값 = row[3].value
                산출식값 = row[4].value
                결과값 = row[5].value

                if 층값 is not None:
                    if '동 명' in 층값:
                        호확정 = 층값.split('-')[-1].strip()
                        실확정 = 층값.split(']')[0].split('[')[-1].replace('-', '').strip()

                        if 실확정 == "부속동":
                            실확정 = "부속동"

                        continue

                if 층값 is not None and 부호값 is not None:
                    if 층값 != "FT":
                        층값 = f'{층값}F'
                    층확정 = 층값
                    부위확정 = 부호값

                if 규격값 is not None:
                    규격_분리_하이픈 = 규격값.split('-')
                    if len(규격_분리_하이픈) == 3:
                         규격값 = f'{규격_분리_하이픈[0]}-{규격_분리_하이픈[1]}-{규격_분리_하이픈[-1].zfill(2)}'

                if '[ 비 고 ]' in 명칭값:
                    continue

                if 명칭값 is not None and 결과값 is not None:

                    명칭값_제거_어퍼스트로피 = 명칭값.replace("'", "")

                    if len(명칭값_제거_어퍼스트로피) >= 2:
                        품명확정 = 명칭값_제거_어퍼스트로피

                    품명확정x = f'{실확정}_{품명확정}'

                    규격확정 = 규격값
                    산식확정 = 산출식값
                    수량확정 = 결과값

                    내역 = ExcelStandard(
                        층=층확정,
                        호=호확정,
                        실=실확정,
                        대공종='건축',
                        중공종='철근콘크리트공사',
                        코드='',
                        품명=품명확정,
                        규격=규격확정,
                        단위='',
                        부위=부위확정,
                        타입='구조',
                        산식=산식확정,
                        수량=수량확정,
                        Remark='',
                        개소=''
                    )
                    산출서목록.append(내역)

    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 15
    new_sheet.column_dimensions["H"].width = 15

    for 내역 in 산출서목록:
        new_sheet.append(내역.to_excel())

    new_workbook.save(saveFilePath)

    # 저장 - 주동A + 기타분류
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    new_sheet.append(head_title)

    saveFileName1 = '구조완성-' + fileCreateDate + '-주동A.xlsx'
    saveFile1 = f'{saveFilePath}{saveFileName1}'

    for 내역 in 산출서목록:
       if (내역.실 == "주동A") or (내역.실 != "주동B" and 내역.실 != '주동C' and 내역.실 != '부동' and 내역.실 != '부속동'):

        new_sheet.append(내역.to_excel())
    new_workbook.save(saveFile1)

    # 저장 - 주동B
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조(데이터변경X)'
    new_sheet.append(head_title)

    saveFileName2 = '구조완성-' + fileCreateDate + '-주동B.xlsx'
    saveFile2 = f'{saveFilePath}{saveFileName2}'
    for 내역 in 산출서목록:
        if 내역.실 == "주동B":
            new_sheet.append(내역.to_excel())
    new_workbook.save(saveFile2)

    # 저장 - 주동C
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조(데이터변경X)'
    new_sheet.append(head_title)

    saveFileName3 = '구조완성-' + fileCreateDate + '-주동C.xlsx'
    saveFile3 = f'{saveFilePath}{saveFileName3}'
    for 내역 in 산출서목록:
       if 내역.실=="주동C":
        new_sheet.append(내역.to_excel())
    new_workbook.save(saveFile3)

    # 저장 - 부동, 부속동
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조(데이터변경X)'
    new_sheet.append(head_title)

    saveFileName4 = '구조완성-' + fileCreateDate + '-부동.xlsx'
    saveFile4 = f'{saveFilePath}{saveFileName4}'
    for 내역 in 산출서목록:
        if 내역.실 == "부동" or 내역.실 == "부속동":
            new_sheet.append(내역.to_excel())
    new_workbook.save(saveFile4)

    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFile1])
        subprocess.call(['open', saveFile2])
        subprocess.call(['open', saveFile3])
        subprocess.call(['open', saveFile4])



if __name__ == '__main__':
    excel_normalize('PyCharm')
