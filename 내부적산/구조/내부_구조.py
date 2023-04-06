from openpyxl import load_workbook, Workbook
from ExcelStandard import ExcelStandard
from datetime import datetime
import platform
import subprocess
import ReplacePersonal

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')
systemOs = platform.system()


# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0316'
##################################


if systemOs == 'Darwin':
    openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/구조.xlsx'
    saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/구조완성-' + fileCreateDate + '.xlsx'
else:
    openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조.xlsx'
    saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)

    산출서목록 = []

    파싱시트목록 = ['부재별산출서', '기타산출서', '아파트옹벽 Unit별산출서']
    print('=================================')
    for 시트명 in 파싱시트목록:
        if 시트명 in excel.sheetnames:
            print("파싱시트체크 : ", 시트명, '(O)')
        else:
            print("파싱시트체크 : ", 시트명, '(X) - 확인필요')
    print('=================================')

    for sheetname in 파싱시트목록:

        if sheetname in excel.sheetnames:
            worksheet = excel[sheetname]

            층확정 = ''
            호확정 = ''
            실확정 = ''
            품명확정 = ''
            규격확정 = ''
            부위확정 = ''
            산식확정 = ''
            수량확정 = ''

            for row in worksheet.iter_rows(min_row=4):

                층값 = str(row[0].value)
                부호값 = row[1].value
                명칭값 = row[2].value
                규격값 = row[3].value
                산출식값 = row[4].value
                결과값 = row[5].value

                if 층값 is not None:
                    if '동 명' in 층값:
                        호확정 = 층값.split('-')[-1].strip()
                        continue

                if 층값 is not None and 부호값 is not None:
                    if 층값 != "FT":
                        층값 = f'{층값}F'
                    층확정 = 층값
                    부위확정 = 부호값

                if 규격값 is not None:
                    규격_분리_하이픈 = 규격값.split('-')
                    if len(규격_분리_하이픈) == 3:
                         규격값 = f'{규격_분리_하이픈[0]}-{규격_분리_하이픈[1]}-{규격_분리_하이픈[-1].zfill(2)}'

                if '[ 비 고 ]' in 명칭값:
                    continue

                if 명칭값 is not None and 결과값 is not None:

                    명칭값_제거_어퍼스트로피 = 명칭값.replace("'", "")

                    if len(명칭값_제거_어퍼스트로피) >= 2:
                        품명확정 = 명칭값_제거_어퍼스트로피

                    규격확정 = 규격값
                    산식확정 = 산출식값
                    수량확정 = 결과값

                    내역 = ExcelStandard(
                        층=층확정,
                        호=호확정,
                        실='',
                        대공종='건축',
                        중공종='철근콘크리트공사',
                        코드='',
                        품명=품명확정,
                        규격=규격확정,
                        단위='',
                        부위=부위확정,
                        타입='구조',
                        산식=산식확정,
                        수량=수량확정,
                        Remark='',
                        개소=''
                    )
                    산출서목록.append(내역)


    # 품명 규격 개인별 지정 변경 S #######################
    for 내역 in 산출서목록:
        ReplacePersonal.launch(내역)
    # 품명 규격 개인별 지정 변경 E #######################


    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 15
    new_sheet.column_dimensions["H"].width = 15

    for 내역 in 산출서목록:
        new_sheet.append(내역.to_excel())

    new_workbook.save(saveFilePath)

    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])
    elif systemOs == "Windows":
        subprocess.Popen(saveFilePath, shell=True)


if __name__ == '__main__':
    excel_normalize('PyCharm')
