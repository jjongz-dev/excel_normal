from openpyxl import load_workbook, Workbook
from ExcelStandard import ExcelStandard
from ExcelGroup import ExcelGroup
from datetime import datetime

import re
import platform
import subprocess
import ReplaceCivilEarthwork
import ReplaceCivilSidePostPile
import ReplaceCivilCIP
import ReplaceCivilStrut
import ReplaceCivilSGR
import ReplaceCivilLW
import ReplaceRaker
import ReplaceCivilRoadDeckingPanel
import ReplaceCivilGlobal
import ReplacePersonal


fileCreateDate = datetime.strftime(datetime.today(), '%y%m%d_%H%M')
systemOs = platform.system()


# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0409'
##################################


if systemOs == 'Darwin':
    openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/토목.xlsx'
    saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/토목완성-' + fileCreateDate + '.xlsx'
else:
    openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\토목.xlsx'
    saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\토목완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath, data_only=True)

    산출서목록 = []
    내역서목록 = []

    파싱시작점기준문자 = ['부위', '도형', '구분', '코드', '품목', '품명']

    파싱시트목록 = ['공종별내역서', '토공집계표', '가시설공 집계표', 'C.I.P 집계표', 'STRUT공 집계표', 'RAKER공 집계표', 'S.G.R공 집계표', 'LW공 집계표', '복공 집계표']
    print('=================================')
    for 시트명 in 파싱시트목록:
        if 시트명 in excel.sheetnames:
            print("파싱시트체크 : ", 시트명, '(O)')
        else:
            print("파싱시트체크 : ", 시트명, '(X) - 확인필요')
    print('=================================')

    품명확정 = ''

    if "토공집계표" in excel.sheetnames:
        worksheet = excel['토공집계표']

        for row in worksheet.iter_rows(min_row=8):

            공종값 = row[0].value
            규격값 = row[8].value
            단위값 = row[15].value
            수량값 = row[19].value

            if 공종값 is not None and 단위값 is not None:
                품명확정 = 공종값.replace('\n', '')

            if 단위값 is None or 수량값 == 0:
                continue

            규격확정 = 규격값
            단위확정 = 단위값
            산식확정 = 수량값
            수량확정 = 수량값

            내역 = ExcelStandard(
                층='',
                호='',
                실='',
                대공종='토목',
                중공종='',
                코드='',
                품명=품명확정,
                규격=규격확정,
                단위=단위확정,
                부위='',
                타입='외부',
                산식=산식확정,
                수량=수량확정,
                Remark='',
                개소=''
            )
            산출서목록.append(내역)

        # 품명 규격 자동 변경 S #######################
        for 내역 in 산출서목록:
            ReplaceCivilEarthwork.launch(내역)
        # 품명 규격 자동 변경 E #######################

    if "가시설공 집계표" in excel.sheetnames:
        worksheet = excel['가시설공 집계표']

        for row in worksheet.iter_rows(min_row=8):

            공종값 = row[1].value
            규격값 = row[9].value
            단위값 = row[16].value
            수량값 = row[20].value
            비고값 = row[25].value

            if 공종값 is not None and 단위값 is not None:
                품명값 = 공종값.replace('\n', '')
                품명확정 = 품명값

            if 품명값 == "H-PILE 연결" and 비고값 is not None:

                품명확정 = f'{품명값}{비고값}'
                #print(품명값, '##', 비고값, '==', 품명확정)

            if 단위값 is None or 수량값 == 0:
                continue

            규격확정 = 규격값
            단위확정 = 단위값
            산식확정 = 수량값
            수량확정 = 수량값

            내역 = ExcelStandard(
                층='',
                호='',
                실='',
                대공종='토목',
                중공종='',
                코드='',
                품명=품명확정,
                규격=규격확정,
                단위=단위확정,
                부위='',
                타입='외부',
                산식=산식확정,
                수량=수량확정,
                Remark='',
                개소=''
            )
            산출서목록.append(내역)

        # 품명 규격 자동 변경 S #######################
        for 내역 in 산출서목록:
            ReplaceCivilSidePostPile.launch(내역)
        # 품명 규격 자동 변경 E #######################

    if "C.I.P 집계표" in excel.sheetnames:
        worksheet = excel['C.I.P 집계표']

        for row in worksheet.iter_rows(min_row=9):

            공종값 = row[1].value
            규격값 = row[9].value
            단위값 = row[16].value
            수량값 = row[20].value
            비고값 = row[25].value

            if 공종값 is not None and 단위값 is not None:
                품명값 = 공종값.replace('\n', '')
                품명확정 = 품명값

            if "CON'C" in 품명값 and 비고값 is not None:
                품명확정 = f'{품명값}{비고값}'
                #print(품명값, '##', 비고값, '==', 품명확정)

            if 단위값 is None or 수량값 == 0:
                continue

            if 규격값 =="풍화암":
                continue

            규격확정 = 규격값
            단위확정 = 단위값
            산식확정 = 수량값
            수량확정 = 수량값

            내역 = ExcelStandard(
                층='',
                호='',
                실='',
                대공종='토목',
                중공종='',
                코드='',
                품명=품명확정,
                규격=규격확정,
                단위=단위확정,
                부위='',
                타입='외부',
                산식=산식확정,
                수량=수량확정,
                Remark='',
                개소=''
            )
            산출서목록.append(내역)

        # 품명 규격 자동 변경 S #######################
        for 내역 in 산출서목록:
            ReplaceCivilCIP.launch(내역)
        # 품명 규격 자동 변경 E #######################

    if "STRUT공 집계표" in excel.sheetnames:
        worksheet = excel['STRUT공 집계표']

        for row in worksheet.iter_rows(min_row=9):

            공종값 = row[1].value
            규격값 = row[9].value
            단위값 = row[16].value
            수량값 = row[20].value
            비고값 = row[25].value

            if 공종값 is not None and 단위값 is not None:
                품명값 = 공종값.replace('\n', '')
                품명확정 = 품명값

            if "H-PILE 연결" in 품명값 and 비고값 is not None:
                품명확정 = f'{품명값}{비고값}'
                #print(품명값, '##', 비고값, '==', 품명확정)

            if 단위값 is None or 수량값 == '#REF!':
                continue

            규격확정 = 규격값
            단위확정 = 단위값
            산식확정 = 수량값
            수량확정 = 수량값

            내역 = ExcelStandard(
                층='',
                호='',
                실='',
                대공종='토목',
                중공종='',
                코드='',
                품명=품명확정,
                규격=규격확정,
                단위=단위확정,
                부위='',
                타입='외부',
                산식=산식확정,
                수량=수량확정,
                Remark='',
                개소=''
            )
            산출서목록.append(내역)

        # 품명 규격 자동 변경 S #######################
        for 내역 in 산출서목록:
            ReplaceCivilStrut.launch(내역)
        # 품명 규격 자동 변경 E #######################

    if "RAKER공 집계표" in excel.sheetnames:
        worksheet = excel['RAKER공 집계표']

        품명확정 =''

        for row in worksheet.iter_rows(min_row=11):

            공종값 = row[1].value
            규격값 = row[9].value
            단위값 = row[16].value
            수량값 = row[20].value


            if 공종값 is not None and 단위값 is not None:
                품명확정 = 공종값

            if 단위값 is None or 수량값 == 0:
                continue


            규격확정 = 규격값
            단위확정 = 단위값
            산식확정 = 수량값
            수량확정 = 수량값

            내역 = ExcelStandard(
                층='',
                호='',
                실='',
                대공종='토목',
                중공종='',
                코드='',
                품명=품명확정,
                규격=규격확정,
                단위=단위확정,
                부위='',
                타입='외부',
                산식=산식확정,
                수량=수량확정,
                Remark='',
                개소=''
            )
            산출서목록.append(내역)

    # 품명 규격 자동 변경 S #######################
        for 내역 in 산출서목록:
            ReplaceRaker.launch(내역)
        # 품명 규격 자동 변경 E #######################


    if "S.G.R공 집계표" in excel.sheetnames:
        worksheet = excel['S.G.R공 집계표']

        품명확정 = ''

        for row in worksheet.iter_rows(min_row=11):

            공종값 = row[1].value
            규격값 = row[9].value
            단위값 = row[16].value
            수량값 = row[20].value

            if 공종값 is not None and 단위값 is not None:
                품명확정 = 공종값

            if 단위값 is None or 수량값 == 0:
                continue

            규격확정 = 규격값
            단위확정 = 단위값
            산식확정 = 수량값
            수량확정 = 수량값

            내역 = ExcelStandard(
                층='',
                호='',
                실='',
                대공종='토목',
                중공종='',
                코드='',
                품명=품명확정,
                규격=규격확정,
                단위=단위확정,
                부위='',
                타입='외부',
                산식=산식확정,
                수량=수량확정,
                Remark='',
                개소=''
            )
            산출서목록.append(내역)

        # 품명 규격 자동 변경 S #######################
        for 내역 in 산출서목록:
            ReplaceCivilSGR.launch(내역)
        # 품명 규격 자동 변경 E #######################

    if "LW공 집계표" in excel.sheetnames:
        worksheet = excel['LW공 집계표']

        품명확정 = ''

        for row in worksheet.iter_rows(min_row=11):

            공종값 = row[1].value
            규격값 = row[9].value
            단위값 = row[16].value
            수량값 = row[20].value

            if 공종값 is not None and 단위값 is not None:
                품명확정 = 공종값

            if 단위값 is None or 수량값 == 0:
                continue

            규격확정 = 규격값
            단위확정 = 단위값
            산식확정 = 수량값
            수량확정 = 수량값

            내역 = ExcelStandard(
                층='',
                호='',
                실='',
                대공종='토목',
                중공종='',
                코드='',
                품명=품명확정,
                규격=규격확정,
                단위=단위확정,
                부위='',
                타입='외부',
                산식=산식확정,
                수량=수량확정,
                Remark='',
                개소=''
            )
            산출서목록.append(내역)

        # 품명 규격 자동 변경 S #######################
        for 내역 in 산출서목록:
            ReplaceCivilLW.launch(내역)
        # 품명 규격 자동 변경 E #######################


    if "복공 집계표" in excel.sheetnames:
        worksheet = excel['복공 집계표']

        품명확정 = ''

        for row in worksheet.iter_rows(min_row=11):

            공종값 = row[1].value
            규격값 = row[9].value
            단위값 = row[16].value
            수량값 = row[20].value

            if 공종값 is not None and 단위값 is not None:
                품명확정 = 공종값

            if 단위값 is None or 수량값 == 0:
                continue

            규격확정 = 규격값
            단위확정 = 단위값
            산식확정 = 수량값
            수량확정 = 수량값

            내역 = ExcelStandard(
                층='',
                호='',
                실='',
                대공종='토목',
                중공종='',
                코드='',
                품명=품명확정,
                규격=규격확정,
                단위=단위확정,
                부위='',
                타입='외부',
                산식=산식확정,
                수량=수량확정,
                Remark='',
                개소=''
            )
            산출서목록.append(내역)

        # 품명 규격 자동 변경 S #######################
        for 내역 in 산출서목록:
            ReplaceCivilRoadDeckingPanel.launch(내역)
        # 품명 규격 자동 변경 E #######################

    # 계측장비6대 추가
    for 넘버링 in range(1, 7):
        품명확정 = f'계측장비#{넘버링}'

        내역 = ExcelStandard(
            층='',
            호='',
            실='',
            대공종='토목',
            중공종='',
            코드='',
            품명=품명확정,
            규격='',
            단위='EA',
            부위='',
            타입='외부',
            산식='★내역서 확인 후 값 변경',
            수량='★내역서 확인 후 값 변경',
            Remark='',
            개소=''
        )
        산출서목록.append(내역)


    # 품명 규격 개인별 지정 변경 S #######################
    for 내역 in 산출서목록:
        ReplacePersonal.launch(내역)
    # 품명 규격 개인별 지정 변경 E #######################



    # 공종별내역서 ###################################################################################

    if '공종별내역서' in excel.sheetnames:

        worksheet = excel['공종별내역서']
        내역서시작줄 = 0
        중공종확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        수량확정 = ''

        for 가로줄번호, row in enumerate(worksheet.rows):

            if row[0].value.replace(' ', '') in 파싱시작점기준문자:
                내역서시작줄 = 가로줄번호 + 3
                break

        for row in worksheet.iter_rows(min_row=내역서시작줄):

            품명값 = row[1].value
            규격값 = row[2].value
            단위값 = row[3].value
            수량값 = row[4].value

            if 품명값 is not None and (단위값 is None or 단위값 == ''):
                if '[ 합           계 ]' in 품명값:
                    continue
                else:
                    중공종 = 품명값.replace(' ', '')
                    중공종확정 = 중공종[re.search('\\d{0,9}', 중공종).end():]
                    continue

            if 품명값 is not None and (단위값 is not None or 단위값 != '') and (수량값 is not None and 수량값 != 0):

                품명확정 = 품명값
                규격확정 = 규격값
                단위확정 = 단위값
                수량확정 = 수량값

                내역 = ExcelGroup(
                    중공종=중공종확정,
                    품명=품명확정,
                    규격=규격확정,
                    단위=단위확정,
                    수량=수량확정
                )
                내역서목록.append(내역)

        # 품명 규격 자동 변경 S #######################
        for 내역 in 내역서목록:
            ReplaceCivilGlobal.launch(내역)
        # 품명 규격 자동 변경 E #######################


    # 엑셀 처리 완료 ##################################################################################





    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '토목(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 40

    for 내역 in 산출서목록:
        new_sheet.append(내역.to_excel())


    sheet = new_workbook.create_sheet(title='집계표')
    sheet.append(['중공종', '품명', '규격', '단위', '수량(할증전)'])
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30

    for 내역 in 내역서목록:
        sheet.append(내역.to_excelGroup())

    new_workbook.save(saveFilePath)

    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()

    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])
    elif systemOs == "Windows":
        subprocess.Popen(saveFilePath, shell=True)


if __name__ == '__main__':
    excel_normalize('PyCharm')
