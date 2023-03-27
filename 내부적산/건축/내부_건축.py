from openpyxl import load_workbook, Workbook
from ExcelStandard import ExcelStandard
from ExcelGroup import ExcelGroup
from collections import defaultdict
from pprint import pprint as pprint
from datetime import datetime

import re
import platform
import subprocess


import ReplaceFinEarthWork
import ReplaceFinDuplicateDelete


fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')

# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0240'
##################################

openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/건축.xlsx'
saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/건축완성-' + fileCreateDate + '.xlsx'


# openFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\건축.xlsx'
# saveFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\건축완성-' + fileCreateDate + '.xlsx'

# openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축.xlsx'
# saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)

    집계표목록 = []
    내역목록 = []

    산출서시작점기준문자 = ['부위', '도형', '구분','코드']

    # 가설산출서 ###################################################################################

    if '가설산출서' in excel.sheetnames:

        worksheet = excel['가설산출서']
        산출서시작줄 = 0
        층확정 = ''
        호확정 = ''
        실확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        부위확정 = ''
        타입확정 = ''
        산식확정 = ''
        수량확정 = ''
        개소값 = ''

        for 가로줄번호, row in enumerate(worksheet.rows):
            if row[0].value in 산출서시작점기준문자:
                산출서시작줄 = 가로줄번호 + 2
                break

        for row in worksheet.iter_rows(min_row=산출서시작줄):

            부위값 = row[0].value
            층범위값 = row[1].value
            품명값 = row[2].value
            규격값 = row[3].value
            단위값 = row[4].value
            산식값 = row[5].value
            층갯수값 = row[6].value
            물량값 = row[7].value

            if (부위값 is not None
                    and (품명값 is None or 품명값 == '')
            ):
                if '개소' in 부위값:
                    호확정 = 실확정 = 부위값.split('개소')[0].split(':')[-1].strip()
                    개소값 = int(부위값.split(':')[-1].strip())
                    continue
                else:
                    continue

            # 층범위값 표기 수정
            if 층범위값 is not None and 층범위값 != '':
                if 'PT' in 층범위값:
                    층범위값 = 'PT'
                elif 'P1' in 층범위값:
                    층범위값 = 'RF'
                else:
                    층범위값 = f'{층범위값}F'

            if 품명값 is not None and (물량값 is not None and 물량값 != 0):

                품명확정 = 품명값
                층확정 = 층범위값
                규격확정 = 규격값
                단위확정 = 단위값
                부위확정 = 부위값
                산식확정 = 산식값
                수량확정 = 물량값
                개소확정 = 개소값

                내역 = ExcelStandard(
                    층=층확정,
                    호=호확정,
                    실=실확정,
                    대공종='건축',
                    중공종='',
                    코드='',
                    품명=품명확정,
                    규격=규격확정,
                    단위=단위확정,
                    부위=부위확정,
                    타입='외부',
                    산식=산식확정,
                    수량=수량확정,
                    Remark='',
                    개소=개소확정
                )
                내역목록.append(내역)

    # 토공산출서 ###################################################################################

    if '토공산출서' in excel.sheetnames:

        worksheet = excel['토공산출서']
        산출서시작줄 = 0
        층확정 = ''
        호확정 = ''
        실확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        부위확정 = ''
        타입확정 = ''
        산식확정 = ''
        수량확정 = ''
        개소값 = ''

        for 가로줄번호, row in enumerate(worksheet.rows):
            if row[0].value in 산출서시작점기준문자:
                산출서시작줄 = 가로줄번호 + 2
                break

        for row in worksheet.iter_rows(min_row=산출서시작줄):

            도형값 = row[0].value
            부위값 = ''
            층범위값 = row[2].value
            품명값 = row[3].value
            규격값 = row[4].value
            단위값 = row[5].value
            산식값 = row[6].value
            층갯수값 = row[7].value
            물량값 = row[8].value

            if (도형값 is not None
                    and (품명값 is None or 품명값 == '')
            ):
                if '개소' in 도형값:
                    호확정 = 실확정 = 도형값.split('개소')[0].split(':')[-1].strip()
                    개소값 = int(도형값.split(':')[-1].strip())
                    continue
                else:
                    continue

            # 층범위값 표기 수정
            if 층범위값 is not None and 층범위값 != '':
                if 'PT' in 층범위값:
                    층범위값 = 'PT'
                elif 'P1' in 층범위값:
                    층범위값 = 'RF'
                else:
                    층범위값 = f'{층범위값}F'

            if 품명값 is not None and (물량값 is not None and 물량값 != 0):

                품명확정 = 품명값
                층확정 = 층범위값
                규격확정 = 규격값
                단위확정 = 단위값
                부위확정 = 부위값
                산식확정 = 산식값
                수량확정 = 물량값
                개소확정 = 개소값

                내역 = ExcelStandard(
                    층='1F',
                    호=호확정,
                    실=실확정,
                    대공종='건축',
                    중공종='',
                    코드='',
                    품명=품명확정,
                    규격=규격확정,
                    단위=단위확정,
                    부위=부위확정,
                    타입='외부',
                    산식=산식확정,
                    수량=수량확정,
                    Remark='',
                    개소=개소확정
                )
                내역목록.append(내역)

        # 품명 규격 자동 변경 S #######################
        for 내역 in 내역목록:
            ReplaceFinEarthWork.launch(내역)
        # 품명 규격 자동 변경 E #######################

    sheet_names = ['내부산출서', '내부산출서-1', '내부산출서-2']
    for sheet in sheet_names:
        if sheet in excel.sheetnames:

            worksheet = excel[sheet]
            산출서시작줄=0
            층확정 = ''
            호확정 = ''
            실확정 = ''
            품명확정 = ''
            규격확정 = ''
            단위확정 = ''
            부위확정 = ''
            타입확정 = ''
            산식확정 = ''
            수량확정 = ''
            개소값 = ''

            for 가로줄번호, row in enumerate(worksheet.rows):

                if row[0].value in 산출서시작점기준문자:
                    산출서시작줄 = 가로줄번호
                    break

            for row in worksheet.iter_rows(min_row=산출서시작줄):

                도형값 = row[0].value
                부위값 = row[1].value
                품명값 = row[2].value
                규격값 = row[3].value
                단위값 = row[4].value
                산식값 = row[5].value
                물량값 = row[6].value

                if 도형값 =='도형' and 부위값 =='부위':
                    continue

                if 도형값 is not None and 품명값 is None and 단위값 is None and 물량값 is None:

                    층값 = 도형값.split(' ')[-2].strip()
                    if '층' in 층값:
                        층확정 = 층값
                        continue
                    if '실명 :' in 도형값:
                        호실값= 도형값.split('개소')[0].split(':')[-1].strip()
                        호실값_분리_호 = 호실값.split('호')
                        if len(호실값_분리_호)>1:
                            호확정 = f'{호실값_분리_호[0].strip()}호'
                            실확정 = 호실값_분리_호[1].strip()
                        else:
                            층확정 = 호실값
                        continue
                    else:
                        continue

                if 품명값 is not None and (물량값 is not None and 물량값 != 0):

                    품명확정 = 품명값
                    규격확정 = 규격값
                    단위확정 = 단위값
                    부위확정 = 부위값
                    산식확정 = 산식값
                    수량확정 = 물량값
                    개소확정 = 개소값

                    내역 = ExcelStandard(
                        층=층확정,
                        호=호확정,
                        실=실확정,
                        대공종='건축',
                        중공종='',
                        코드='',
                        품명=품명확정,
                        규격=규격확정,
                        단위=단위확정,
                        부위='',
                        타입='내부',
                        산식=산식확정,
                        수량=수량확정,
                        Remark='',
                        개소=개소확정
                    )
                    내역목록.append(내역)

    # 외부산출서 ###################################################################################

    if '외부산출서' in excel.sheetnames:

        worksheet = excel['외부산출서']
        산출서시작줄=0
        층확정 = ''
        호확정 = ''
        실확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        부위확정 = ''
        타입확정 = ''
        산식확정 = ''
        수량확정 = ''
        개소값 = ''

        for 가로줄번호, row in enumerate(worksheet.rows):
            if row[0].value in 산출서시작점기준문자:
                산출서시작줄 = 가로줄번호 + 2
                break

        for row in worksheet.iter_rows(min_row=산출서시작줄):

            도형값 = row[0].value
            부위값 = ''
            층범위값 = row[2].value
            품명값 = row[3].value
            규격값 = row[4].value
            단위값 = row[5].value
            산식값 = row[6].value
            층갯수값 = row[7].value
            물량값 = row[8].value

            if 도형값 is not None and 품명값 is None and 단위값 is None and 물량값 is None:

                if '구분명 :' in 도형값:
                    층호실 = 도형값.split('개소')[0].split(':')[-1].strip()
                    층호실_분리_언더바 = 층호실.split('_')
                    층호실_분리_띄어쓰기 = 층호실.split(' ')

                    if len(층호실_분리_언더바)>1:
                        층확정 = 층호실_분리_언더바[0].strip()
                        호확정 = 실확정 = 층호실_분리_언더바[-1].strip()
                    elif len(층호실_분리_띄어쓰기) > 1:
                        층확정 = 층호실_분리_띄어쓰기[0].strip()
                        호확정 = 실확정 = 층호실_분리_띄어쓰기[-1].strip()
                    else:
                        층확정 = 층호실

                    개소값 = int(도형값.split(':')[-1].strip())
                    continue
                else:
                    continue

            # 층범위에 값이 있는경우 그값을 층정보로 사용
            if 층범위값 is not None:
                if 'P1' in 층범위값:
                    층확정 = 'RF'
                else:
                    if re.match('\\d{1,2}', 층범위값):
                        층확정 = f'{층범위값}F'

            if 물량값 =="'" or 물량값 =='0' or 물량값 ==0 or 물량값 =='물량':
                continue

            if 품명값 is not None and (물량값 is not None and 물량값 != 0):

                품명확정 = 품명값
                규격확정 = 규격값
                단위확정 = 단위값
                부위확정 = 부위값
                산식확정 = 산식값
                수량확정 = 물량값
                개소확정 = 개소값

                내역 = ExcelStandard(
                    층=층확정,
                    호=호확정,
                    실=실확정,
                    대공종='건축',
                    중공종='',
                    코드='',
                    품명=품명확정,
                    규격=규격확정,
                    단위=단위확정,
                    부위=부위확정,
                    타입='외부',
                    산식=산식확정,
                    수량=수량확정,
                    Remark='',
                    개소=개소확정
                )
                내역목록.append(내역)

    # 철골산출서 ###################################################################################
    if '철골산출서' in excel.sheetnames:

        worksheet = excel['철골산출서']
        산출서시작줄 =0
        층확정 = ''
        호확정 = ''
        실확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        부위확정 = ''
        타입확정 = ''
        산식확정 = ''
        수량확정 = ''
        개소값 = ''

        for 가로줄번호, row in enumerate(worksheet.rows):
            if row[0].value in 산출서시작점기준문자:
                산출서시작줄 = 가로줄번호 + 2
                break

        for row in worksheet.iter_rows(min_row=산출서시작줄):

            부위값 = row[0].value
            층범위값 = row[1].value
            품명값 = row[2].value
            규격값 = row[3].value
            단위값 = row[4].value
            산식값 = row[5].value
            층갯수값 = row[6].value
            물량값 = row[7].value

            if (부위값 is not None
                    and (품명값 is None or 품명값 == '')
            ):
                if '개소' in 부위값:
                    호확정 = 실확정 = 부위값.split('개소')[0].split(':')[-1].strip()
                    개소값 = int(부위값.split(':')[-1].strip())
                    continue
                else:
                    continue

            if 품명값 is not None and (물량값 is not None and 물량값 != 0):

                품명확정 = 품명값
                층확정 = 층범위값
                규격확정 = 규격값
                단위확정 = 단위값
                부위확정 = 부위값
                산식확정 = 산식값
                수량확정 = 물량값
                개소확정 = 개소값

                내역 = ExcelStandard(
                    층=층확정,
                    호=호확정,
                    실=실확정,
                    대공종='건축',
                    중공종='',
                    코드='',
                    품명=품명확정,
                    규격=규격확정,
                    단위=단위확정,
                    부위=부위확정,
                    타입='내부',
                    산식=산식확정,
                    수량=수량확정,
                    Remark='',
                    개소=개소확정
                )
                내역목록.append(내역)

    # 동별창호리스트 ###################################################################################
    if '동별창호리스트' in excel.sheetnames:
        worksheet = excel['동별창호리스트']

        창호층목록 = defaultdict(str)
        창호별층별수량 = defaultdict(list)
        산출서시작줄 = 0
        구분시작줄 = 0
        층확정 = ''
        호확정 = ''
        실확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        부위확정 = ''
        타입확정 = ''
        산식확정 = ''
        수량확정 = ''
        개소값 = ''

        for 가로줄번호, row in enumerate(worksheet.rows):
            if row[0].value in 산출서시작점기준문자:
                구분시작줄 = 가로줄번호
                산출서시작줄 = 가로줄번호 + 2
                break

        for 세로줄번호, 층정보 in enumerate(list(worksheet.rows)[구분시작줄]):
            #print(층정보.value)
            if re.match('[BFP]\\d{1,2}', 층정보.value):
                창호층목록[층정보.value] = 세로줄번호

        for row in worksheet.iter_rows(min_row=산출서시작줄):

            구분값 = row[0].value
            창호명값 = row[1].value
            가로값 = row[2].value
            세로값 = row[3].value
            면적값 = row[4].value
            산식값 = row[5].value
            합계값 = row[10].value

            if 창호명값 is not None:

                #창호층수 만큼 반복하며 창호 수량이 있으면 창호를 생성
                for 창호층 in 창호층목록:

                    창호수량 = row[창호층목록[창호층]].value

                    if 창호수량 is not None:
                        if 'B' in 창호층:
                            층값 = f'{창호층}F'
                        elif 'F' in 창호층:
                            창호층수정 = 창호층.replace('F', '')
                            층값 = f'{창호층수정}F'
                        elif 'p1' in 창호층:
                            층값 = 'RF'
                        elif 'p2' in 창호층:
                            층값 = 'PHRF'
                        else:
                            층값 = 창호층

                        층확정 = 층값
                        호확정 = 창호명값
                        품명확정 = 창호명값
                        규격확정 = f'{가로값:0.3f}*{세로값:0.3f}={면적값:0.3f}'
                        수량확정 = row[창호층목록[창호층]].value
                        산식확정 = 수량확정
                        개소확정 = 합계값

                        창호별층별수량[창호명값].append((층확정, 수량확정, 개소확정))

                        내역 = ExcelStandard(
                            층=층확정,
                            호=호확정,
                            실=실확정,
                            대공종='건축',
                            중공종='',
                            코드='',
                            품명=품명확정,
                            규격=규격확정,
                            단위='EA',
                            부위='',
                            타입='창호',
                            산식=산식확정,
                            수량=수량확정,
                            Remark='',
                            개소=개소확정
                        )
                        내역목록.append(내역)

    # 창호산출서 ###################################################################################
    if '창호산출서' in excel.sheetnames:

        #pprint(창호별층별수량)

        worksheet = excel['창호산출서']
        산출서시작줄 = 0
        층확정 = ''
        호확정 = ''
        실확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        부위확정 = ''
        타입확정 = ''
        산식확정 = ''
        수량확정 = ''
        개소값 = ''

        for 가로줄번호, row in enumerate(worksheet.rows):
            if row[0].value in 산출서시작점기준문자:
                산출서시작줄 = 가로줄번호+2
                break

        for row in worksheet.iter_rows(min_row=산출서시작줄):

            부위값 = row[0].value
            품명값 = row[1].value
            규격값 = row[2].value
            단위값 = row[3].value
            산식값 = row[4].value
            물량값 = row[5].value

            if (부위값 is not None
                    and (품명값 is None or 품명값 == '')
            ):
                if '창호명' in 부위값:
                    창호명 = 호확정 = 실확정 = 부위값.split('(')[0].split(':')[-1].strip()
                    continue
                else:
                    continue

            if 물량값 == "'" or 물량값 == '0' or 물량값 == 0 or 물량값 == '물량':
                continue

            if 품명값 is not None and 단위값 is not None and 물량값 is not None:
                #pprint(창호별층별수량[창호명])

                # 창호층수 만큼 반복하며 창호 수량이 있으면 창호를 생성
                for 창호층, 창호수량, 개소값 in 창호별층별수량[창호명]:

                    층확정 = 창호층
                    호확정 = 창호명
                    품명확정 = 품명값
                    규격확정 = 규격값
                    단위확정 = 단위값
                    산식확정 = f'({산식값})*<수량>({창호수량})'
                    수량확정 = float(물량값) * float(창호수량)
                    개소확정 = 개소값

                    내역 = ExcelStandard(
                        층=층확정,
                        호=호확정,
                        실=실확정,
                        대공종='건축',
                        중공종='',
                        코드='',
                        품명=품명확정,
                        규격=규격확정,
                        단위=단위확정,
                        부위='',
                        타입='창호',
                        산식=산식확정,
                        수량=수량확정,
                        Remark='',
                        개소=개소확정
                    )
                    내역목록.append(내역)

    # 품명 규격 자동 변경 S #######################
    for 내역 in 내역목록:
        ReplaceFinDuplicateDelete.launch(내역)
    # 품명 규격 자동 변경 E #######################




    # 가설산출서 ###################################################################################

    if '공종별집계표' in excel.sheetnames:

        worksheet = excel['공종별집계표']
        산출서시작줄 = 0
        중공종확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        수량확정 = ''

        for 가로줄번호, row in enumerate(worksheet.rows):
            if row[0].value in 산출서시작점기준문자:
                산출서시작줄 = 가로줄번호 + 2
                break

        for row in worksheet.iter_rows(min_row=산출서시작줄):

            품명값 = row[1].value
            규격값 = row[2].value
            단위값 = row[3].value
            물량값 = row[4].value

            if 품명값 is not None and 물량값 is None:
                중공종확정 = 품명값


            if 품명값 is not None and (물량값 is not None and 물량값 != 0):

                품명확정 = 품명값
                규격확정 = 규격값
                단위확정 = 단위값
                수량확정 = 물량값

                집계표내역 = ExcelGroup(
                    중공종=중공종확정,
                    품명=품명확정,
                    규격=규격확정,
                    단위=단위확정,
                    수량=수량확정
                )
                집계표목록.append(집계표내역)

    # 토공산출서 ###################################################################################




    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark', '개소(확인용)']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30

    for 내역 in 내역목록:
       new_sheet.append(내역.to_excel())



    sheet = new_workbook.create_sheet(title='집계표')
    sheet.append(['중공종', '품명', '규격', '단위', '수량(할증전)'])
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30
    for 집계표내역 in 집계표목록:
        sheet.append(집계표내역.to_excelGroup())


    new_workbook.save(saveFilePath)

    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])


if __name__ == '__main__':
    excel_normalize('PyCharm')

