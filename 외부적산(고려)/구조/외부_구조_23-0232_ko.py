from openpyxl import load_workbook, Workbook
from ExcelStandard import ExcelStandard
from datetime import datetime
from collections import defaultdict

from pprint import pprint as pp

import platform
import subprocess

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')

# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0232_ko'
##################################

openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/구조.xlsx'
saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/구조완성-' + fileCreateDate + '.xlsx'

# openFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\구조.xlsx'
# saveFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\구조완성-' + fileCreateDate + '.xlsx'

# openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조.xlsx'
# saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)

    내역목록 = []
    철근이음길이수량목록 = defaultdict(float)
    철근수량목록 = defaultdict(float)
    철근이음길이합산기준수량 = defaultdict(float)

    if '이음길이' in excel.sheetnames:
        worksheet = excel['이음길이']

        for row in worksheet.rows:

            품명값 = row[0].value
            철근이음길이값 = row[1].value
            철근이음길이수량목록[품명값] = 철근이음길이값

    if '부재별산출서' in excel.sheetnames:
        worksheet = excel['부재별산출서']

        for row in worksheet.iter_rows(min_row=5):

            규격값 = row[3].value
            결과값 = row[5].value

            if 규격값 is not None and 결과값 is not None:
                철근수량목록[규격값] = 철근수량목록[규격값] + 결과값

        for 규격 in 철근이음길이수량목록:
            철근이음길이합산기준수량[규격] = 철근이음길이수량목록[규격] / 철근수량목록[규격] * 1000000

    sheetnames = ['부재별산출서']

    for sheetname in sheetnames:

        if sheetname in excel.sheetnames:
            worksheet = excel[sheetname]

            층확정 = ''
            호확정 = ''
            실확정 = ''
            품명확정 = ''
            규격확정 = ''
            부위확정 = ''
            산식확정 = ''
            수량확정 = ''

            for row in worksheet.iter_rows(min_row=4):

                층값 = str(row[0].value)
                부호값 = row[1].value
                명칭값 = row[2].value
                규격값 = row[3].value
                산출식값 = row[4].value
                결과값 = row[5].value

                if 층값 is not None:
                    if '동 명' in 층값:
                        호확정 = 층값.split('-')[-1].strip()
                        continue

                if 층값 is not None and 부호값 is not None:
                    층확정 = 층값
                    부위확정 = 부호값

                    if 'FT' in 층확정:
                        층확정 = 'FT'
                    elif 'PH1' in 층확정:
                        층확정 = 'RF'
                    else:
                        층확정 = f'{층확정}F'

                if 규격값 is not None:
                    규격_분리_하이픈 = 규격값.split('-')
                    if len(규격_분리_하이픈) == 3:
                         규격값 = f'{규격_분리_하이픈[0]}-{규격_분리_하이픈[1]}-{규격_분리_하이픈[-1].zfill(2)}'

                if 명칭값 is not None and 결과값 is not None:

                    명칭값_제거_어퍼스트로피 = 명칭값.replace("'", "")

                    if len(명칭값_제거_어퍼스트로피) >= 2:
                        품명확정 = 명칭값_제거_어퍼스트로피

                    if 규격값 in 철근이음길이수량목록:
                        결과값 = 결과값 + (결과값*철근이음길이합산기준수량[규격값]/1000000)

                    규격확정 = 규격값
                    산식확정 = 산출식값
                    수량확정 = 결과값

                    내역 = ExcelStandard(
                        층=층확정,
                        호=호확정,
                        실='',
                        대공종='건축',
                        중공종='철근콘크리트공사',
                        코드='',
                        품명=품명확정,
                        규격=규격확정,
                        단위='',
                        부위=부위확정,
                        타입='구조',
                        산식=산식확정,
                        수량=수량확정,
                        Remark='',
                        개소=''
                    )
                    내역목록.append(내역)

    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조완성'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 15
    new_sheet.column_dimensions["H"].width = 15

    for 내역 in 내역목록:
        new_sheet.append(내역.to_excel())

    new_workbook.save(saveFilePath)

    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])


if __name__ == '__main__':
    excel_normalize('PyCharm')
