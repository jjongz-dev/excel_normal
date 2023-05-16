from openpyxl import load_workbook, Workbook
from ExcelStandard import ExcelStandard
from ExcelGroup import ExcelGroup
from datetime import datetime

import re
import platform
import subprocess
import ReplacePersonal

fileCreateDate = datetime.strftime(datetime.today(), '%y%m%d_%H%M')
systemOs = platform.system()


# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0361_ko'
##################################


if systemOs == 'Darwin':
    openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/구조.xlsx'
    saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/end-2-gujo-' + fileCreateDate + '.xlsx'
else:
    openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조.xlsx'
    saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\구조완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath)

    산출서목록 = []
    내역서목록 = []
    파싱시작점기준문자 = ['부호', '품명', '품목']

    파싱시트목록 = ['부재별산출서', '기타산출서', '아파트옹벽 Unit별산출서']
    print('=================================')
    for 시트명 in 파싱시트목록:
        if 시트명 in excel.sheetnames:
            print("파싱시트체크 : ", 시트명, '(O)')
        else:
            print("파싱시트체크 : ", 시트명, '(X) - 확인필요')

    if "공종별내역서" in excel.sheetnames:
        print("파싱시트체크 : ", "공종별내역서", '(O)')
    else:
        print("파싱시트체크 : ", "공종별내역서", '(X) - 확인필요')

    print('=================================')

    for sheetname in 파싱시트목록:

        if sheetname in excel.sheetnames:
            worksheet = excel[sheetname]

            층확정 = ''
            호확정 = ''
            실확정 = ''
            품명확정 = ''
            규격확정 = ''
            부위확정 = ''
            산식확정 = ''
            수량확정 = ''

            for 가로줄번호, row in enumerate(worksheet.rows):
                if row[1].value in 파싱시작점기준문자:
                    산출서시작줄 = 가로줄번호 + 2
                    break

            for row in worksheet.iter_rows(min_row=산출서시작줄):

                층값 = str(row[0].value)
                부호값 = row[1].value
                명칭값 = row[2].value
                규격값 = row[3].value
                산출식값 = row[4].value
                결과값 = row[5].value

                if 층값 is not None:
                    if '동 명' in 층값:
                        호확정 = 층값.split('-')[-1].strip()
                        continue

                if 층값 is not None and 부호값 is not None:
                    if 층값 != "FT":
                        층값 = f'{층값}F'
                    층확정 = 층값
                    부위확정 = 부호값


                층확정원본값 = 층확정
                층별데이터복사 = False

                층목록 = []

                if 층확정 is not None:

                    # 2~5F
                    지상층패턴 = r'(\d)[~](\d)[F]'
                    지상층검색결과 = re.search(지상층패턴, 층확정)


                if 지상층검색결과 is not None:

                    층별데이터복사 = True

                    시작층 = int(지상층검색결과.group(1))
                    종료층 = int(지상층검색결과.group(2)) + 1

                    if 시작층 >= 종료층:
                        시작층 = int(지상층검색결과.group(2))
                        종료층 = int(지상층검색결과.group(1)) + 1

                    for 층 in range(시작층, 종료층):
                        층값 = f'{층}F'
                        층목록.append(층값)

                if 규격값 is not None:
                    규격_분리_하이픈 = 규격값.split('-')
                    if len(규격_분리_하이픈) == 3:
                         규격값 = f'{규격_분리_하이픈[0]}-{규격_분리_하이픈[1]}-{규격_분리_하이픈[-1].zfill(2)}'

                if '[ 비 고 ]' in 명칭값:
                    continue






                if 명칭값 is not None and 결과값 is not None:

                    명칭값_제거_어퍼스트로피 = 명칭값.replace("'", "")

                    if len(명칭값_제거_어퍼스트로피) >= 2:
                        품명확정 = 명칭값_제거_어퍼스트로피

                    규격확정 = 규격값
                    산식확정 = 산출식값
                    수량확정 = 결과값

                    if 층별데이터복사:

                        원본수량 = 수량확정
                        분할수량 = 0
                        분할수량합계 = 0
                        총층수 = len(층목록)

                        for index, 층 in enumerate(층목록):

                            층변환 = str(층).zfill(2)
                            층변환확정 = 층

                            산식변환확정 = 산식확정[:산식확정.rfind('*')]
                            분할수량 = round(수량확정/총층수, 3)

                            if index == 총층수 - 1:
                                분할수량 = 원본수량 - 분할수량합계
                            else:
                                분할수량합계 = 분할수량 + 분할수량합계

                            내역 = ExcelStandard(
                                층=층변환확정,
                                호=호확정,
                                실='',
                                대공종='건축',
                                중공종='철근콘크리트공사',
                                코드='',
                                품명=품명확정,
                                규격=규격확정,
                                단위='',
                                부위=부위확정,
                                타입='구조',
                                산식=산식확정,
                                수량=분할수량,
                                Remark='',
                                개소=''
                            )
                            산출서목록.append(내역)


                    else:

                        내역 = ExcelStandard(
                            층=층확정,
                            호=호확정,
                            실='',
                            대공종='건축',
                            중공종='철근콘크리트공사',
                            코드='',
                            품명=품명확정,
                            규격=규격확정,
                            단위='',
                            부위=부위확정,
                            타입='구조',
                            산식=산식확정,
                            수량=수량확정,
                            Remark='',
                            개소=''
                        )
                        산출서목록.append(내역)




    # 품명 규격 개인별 지정 변경 S #######################
    for 내역 in 산출서목록:
        ReplacePersonal.launch(내역)
    # 품명 규격 개인별 지정 변경 E #######################



    # 공종별내역서 ###################################################################################

    if '공종별내역서' in excel.sheetnames:

        내역서집계표중공종 = ['철근콘크리트공사']

        worksheet = excel['공종별내역서']
        내역서시작줄 = 0
        중공종확정 = ''
        품명확정 = ''
        규격확정 = ''
        단위확정 = ''
        수량확정 = ''

        for 가로줄번호, row in enumerate(worksheet.rows):

            #print(row[0].value)

            if row[0].value.replace(' ', '') in 파싱시작점기준문자:
                내역서시작줄 = 가로줄번호 + 3
                break

        for row in worksheet.iter_rows(min_row=내역서시작줄):

            품명값 = row[0].value
            규격값 = row[1].value
            단위값 = row[2].value
            수량값 = row[3].value

            if 품명값 is not None and (단위값 is None or 단위값 == ''):
                if '[ 합           계 ]' in 품명값:
                    continue
                else:
                    중공종 = 품명값.replace(' ', '')
                    중공종확정 = 중공종[re.search('\\d{0,9}', 중공종).end():]
                    continue

            if 중공종확정 in 내역서집계표중공종 and 품명값 is not None and (단위값 is not None or 단위값 != '') and (수량값 is not None and 수량값 != 0):

                품명확정 = 품명값
                규격확정 = 규격값
                단위확정 = 단위값
                수량확정 = 수량값

                내역 = ExcelGroup(
                    중공종=중공종확정,
                    품명=품명확정,
                    규격=규격확정,
                    단위=단위확정,
                    수량=수량확정
                )
                내역서목록.append(내역)


    # 엑셀 처리 완료 ###################################################################################


    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '구조(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 15
    new_sheet.column_dimensions["H"].width = 15

    for 내역 in 산출서목록:
        new_sheet.append(내역.to_excel())



    sheet = new_workbook.create_sheet(title='집계표')
    sheet.append(['중공종', '품명', '규격', '단위', '수량(할증전)'])
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30

    for 내역 in 내역서목록:
        sheet.append(내역.to_excelGroup())





    new_workbook.save(saveFilePath)

    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])
    elif systemOs == "Windows":
        subprocess.Popen(saveFilePath, shell=True)


if __name__ == '__main__':
    excel_normalize('PyCharm')
