
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from Architect.FIN.ItemStandard import ItemStandard
from Architect.FIN.ItemStandard2 import ItemStandard2
from Architect.FIN.PasingRule import earthwork, internal_construction, delete_duplicate
from datetime import datetime
import re
import platform
import subprocess

fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')


# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0275_ko'
##################################

openFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/건축.xlsx'
saveFilePath = '/Users/blue/hb/quantity/'+siteTicketNo+'/건축완성-' + fileCreateDate + '.xlsx'

#openFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\건축.xlsx'
#saveFilePath = 'C:\\howbuild\\quantity\\'+siteTicketNo+'\건축완성-' + fileCreateDate + '.xlsx'

#openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축.xlsx'
#saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\건축완성-' + fileCreateDate + '.xlsx'


def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath, data_only=True)

    items = []
    earthwork_row_index_list = []
    internal_row_index_list = []
    windows_dict = defaultdict(list)

    start_titles = ['부위', '도형', '구분']

    # temporary_constrution
    if '가설산출서' in excel.sheetnames:
        worksheet = excel['가설산출서']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    title_row_index = row_index
                    break
            row_index += 1

        for index, column in enumerate(worksheet[title_row_index]):
            match column.value:
                case '부위':
                    part_index = index
                case '품명':
                    name_index = index
                case '규격':
                    standard_index = index
                case '단위':
                    unit_index = index
                case '산식':
                    formular_index = index
                case '물량':
                    quantity_index = index
                case '층범위':
                    rangefloor_index = index

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=(title_row_index+1)):
            if (row[part_index].value is not None
                    and row[name_index].value is None
                    and row[standard_index].value is None
                    and row[unit_index].value is None
                    and row[formular_index].value is None
                    and row[quantity_index].value is None
            ):
                try:
                    if '개소' in row[part_index].value:
                        temp_split_roomname = row[part_index].value.split('개소')[0]
                        if '구분명' in temp_split_roomname:
                            temp_roomname = temp_split_roomname.split(':')[-1].strip(' ')
                        continue
                except Exception as e:
                    print('예외가 발생했습니다.', e)
                    print('가설산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[quantity_index].value is None
                    or row[quantity_index].value == "'"
                    or row[quantity_index].value == '0'
                    or row[quantity_index].value == 0
            ):
                continue

            if row[rangefloor_index].value is not None:
                temp_rangefloor = row[rangefloor_index].value
                if 'FT' in temp_rangefloor:
                    temp_rangefloor = 'FT'
                elif 'P1' in temp_rangefloor:
                    temp_rangefloor = 'RF'
                else:
                    if re.match('\\d{1,2}', temp_rangefloor):
                        temp_rangefloor = temp_rangefloor + 'F'

            item = ItemStandard(
                floor=temp_rangefloor,
                location=temp_roomname,
                roomname=temp_roomname,
                name=row[name_index].value,
                standard=row[standard_index].value,
                unit=row[unit_index].value,
                type='외부',
                formula=row[formular_index].value,
                sum=row[quantity_index].value,
            )
            items.append(item)

    # earthwork
    if '토공산출서' in excel.sheetnames:
        worksheet = excel['토공산출서']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    title_row_index = row_index
                    earthwork_row_index_list.append(title_row_index)
                    break
            row_index += 1

        for index, column in enumerate(worksheet[earthwork_row_index_list[0]]):
            match column.value:
                case '도형':
                    figure_index = index
                case '품명':
                    name_index = index
                case '규격':
                    standard_index = index
                case '단위':
                    unit_index = index
                case '산식':
                    formular_index = index
                case '물량':
                    quantity_index = index

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row,
                                       min_row=(earthwork_row_index_list[0] + 1)):
            if (row[figure_index].value is not None
                    and row[name_index].value is None
                    and row[standard_index].value is None
                    and row[unit_index].value is None
                    and row[formular_index].value is None
                    and row[quantity_index].value is None
            ):
                try:
                    temp_split_roomname = row[figure_index].value.split('개소')[0]
                    if '구분명' in temp_split_roomname:
                        temp_roomname = temp_split_roomname.split(':')[-1].strip('[] ')
                    continue
                except Exception as e:
                    print('예외가 발생했습니다.', e)
                    print('토공산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[quantity_index].value is None
                    or row[quantity_index].value == "'"
                    or row[quantity_index].value == '0'
                    or row[quantity_index].value == 0
            ):
                continue

            item = ItemStandard(
                floor='1F',
                location=temp_roomname,
                roomname=temp_roomname,
                name=row[name_index].value,
                standard=row[standard_index].value,
                unit=row[unit_index].value,
                type='외부',
                formula=row[formular_index].value,
                sum=row[quantity_index].value,
            )
            items.append(item)

        for item in items:
            earthwork.launch(item)

    # 시트 추가가 필요한경우 오른쪽과 같이 추가 :   ,'시트명'
    # 추가해놓은 시트가 없는경우 자동으로 다음시트로 넘어감.
    sheet_names = ['내부산출서','내부산출서-1','내부산출서-2']
    for sheet in sheet_names:
        # internal_construction
        if sheet in excel.sheetnames:
            worksheet = excel[sheet]
            row_index = 1
            for row in worksheet.iter_rows():
                for column_index, column in enumerate(row):
                    if column.value in start_titles:
                        internal_row_index_list.append(row_index)
                        break
                row_index += 1

            for index, column in enumerate(worksheet[internal_row_index_list[0]]):
                match column.value:
                    case '도형':
                        figure_index = index
                    case '품명':
                        name_index = index
                    case '규격':
                        standard_index = index
                    case '단위':
                        unit_index = index
                    case '산식':
                        formular_index = index
                    case '물량':
                        quantity_index = index

            temp_levels = ""
            temp_roomname = ""
            for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row,
                                           min_row=(internal_row_index_list[0] - 1)):
                if (row[figure_index].value is not None
                        and row[name_index].value is None
                        and row[standard_index].value is None
                        and row[unit_index].value is None
                        and row[formular_index].value is None
                        and row[quantity_index].value is None
                ):
                    try:
                        temp_split_levels = row[figure_index].value.split(' ')[-2]
                        if '층' in temp_split_levels:
                            temp_levels = temp_split_levels
                        temp_split_roomname = row[figure_index].value.split('개소')[0]
                        if '실명' in temp_split_roomname:
                            temp_roomname = temp_split_roomname.split(':')[-1].strip('[] ')
                        continue
                    except Exception as e:
                        print('예외가 발생했습니다.', e)
                        print('내부산출서 : split오류' + str(item))

                # 품명없음 삭제
                if (row[quantity_index].value is None
                        or row[quantity_index].value == "'"
                        or row[quantity_index].value == '0'
                        or row[quantity_index].value == 0
                        or row[quantity_index].value == '물량'
                ):
                    continue

                item = ItemStandard(
                    floor=temp_levels,
                    location=temp_roomname,
                    roomname=temp_roomname,
                    name=row[name_index].value,
                    standard=row[standard_index].value,
                    unit=row[unit_index].value,
                    type='내부',
                    formula=row[formular_index].value,
                    sum=row[quantity_index].value,
                )
                items.append(item)

            for item in items:
                internal_construction.launch(item)


    # external_construction
    if '외부산출서' in excel.sheetnames:
        worksheet = excel['외부산출서']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    external_row_index = row_index
                    break
            row_index += 1

        for index, column in enumerate(worksheet[external_row_index]):
            match column.value:
                case '도형':
                    figure_index = index
                case '품명':
                    name_index = index
                case '규격':
                    standard_index = index
                case '단위':
                    unit_index = index
                case '산식':
                    formular_index = index
                case '물량':
                    quantity_index = index
                case '층범위':
                    rangefloor_index = index

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=external_row_index+1):
            if (row[figure_index].value is not None
                    and row[name_index].value is None
                    and row[standard_index].value is None
                    and row[unit_index].value is None
                    and row[formular_index].value is None
                    and row[quantity_index].value is None
            ):
                try:
                    temp_split_roomname = row[figure_index].value.split('개소')[0]
                    if '구분명' in temp_split_roomname:
                        temp_roomname = temp_split_roomname.split(':')[-1].strip('[] ')
                        if '_' in temp_roomname:
                            levels = temp_roomname.split('_')[0]
                            roomname = temp_roomname.split('_')[-1]
                        else:
                            levels = temp_roomname
                            roomname = temp_roomname
                            continue
                except Exception as e:
                    print('예외가 발생했습니다', e)
                    print('외부산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[quantity_index].value is None
                    or row[quantity_index].value == "'"
                    or row[quantity_index].value == '0'
                    or row[quantity_index].value == 0
                    or row[quantity_index].value == '물량'
            ):
                continue

            if row[rangefloor_index].value is not None:
                temp_rangefloor = row[rangefloor_index].value
                if 'P1' in temp_rangefloor:
                    levels = 'RF'
                else:
                    if re.match('\\d{1,2}', temp_rangefloor):
                        levels = temp_rangefloor + 'F'

            item = ItemStandard(
                floor=levels,
                location=roomname,
                roomname=roomname,
                name=row[name_index].value,
                standard=row[standard_index].value,
                unit=row[unit_index].value,
                type='외부',
                formula=row[formular_index].value,
                sum=row[quantity_index].value,
            )
            items.append(item)

    # external_construction
    if '철골산출서' in excel.sheetnames:
        worksheet = excel['철골산출서']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    external_row_index = row_index
                    break
            row_index += 1

        for index, column in enumerate(worksheet[external_row_index]):
            match column.value:
                case '부위':
                    figure_index = index
                case '품명':
                    name_index = index
                case '규격':
                    standard_index = index
                case '단위':
                    unit_index = index
                case '산식':
                    formular_index = index
                case '물량':
                    quantity_index = index
                case '층범위':
                    rangefloor_index = index

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=external_row_index+1):
            if (row[figure_index].value is not None
                    and row[name_index].value is None
                    and row[standard_index].value is None
                    and row[unit_index].value is None
                    and row[formular_index].value is None
                    and row[quantity_index].value is None
            ):
                try:
                    temp_split_roomname = row[figure_index].value.split('개소')[0]
                    if '구분명' in temp_split_roomname:
                        temp_split2_roomname = temp_split_roomname.split(':')[-1].strip('[] ')
                        continue
                except Exception as e:
                    print('예외가 발생했습니다', e)
                    print('외부산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[quantity_index].value is None
                    or row[quantity_index].value == "'"
                    or row[quantity_index].value == '0'
                    or row[quantity_index].value == 0
                    or row[quantity_index].value == '물량'
            ):
                continue

            if row[rangefloor_index].value is not None:
                temp_rangefloor = row[rangefloor_index].value

            item = ItemStandard(
                floor=temp_rangefloor,
                location=temp_split2_roomname,
                roomname=temp_split2_roomname,
                name=row[name_index].value,
                standard=row[standard_index].value,
                unit=row[unit_index].value,
                type='내부',
                formula=row[formular_index].value,
                sum=row[quantity_index].value,
            )
            items.append(item)

    # List of windows and doors by Dong
    if '동별창호리스트' in excel.sheetnames:
        worksheet = excel['동별창호리스트']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    dong_windowslist_row_index = row_index
                    break
            row_index += 1

        for index, column in enumerate(worksheet[dong_windowslist_row_index]):
            match column.value:
                case '창호명':
                    name_index = index
                case '가로':
                    width_index = index
                case '세로':
                    height_index = index
                case '면적':
                    area_index = index
                case '비고':
                    note_index = index
                case '합계':
                    quantity_index = index

        title_list = []
        for cell_obj in list(worksheet.rows)[dong_windowslist_row_index-1]:

            print('cell_obj.value : ', cell_obj.value)

            if cell_obj.value is not None:
                title_list.append(cell_obj.value)
                # print(title_list)

        floor_list = []
        for title in title_list:
            if re.match('[BFP]\\d{1,2}', title):
                floor_list.append(title)
                # print(floor_list)

        for row in worksheet.iter_rows(min_col=0, min_row=dong_windowslist_row_index+1):
            window_name = row[title_list.index('창호명')].value
            print(window_name)
            if window_name is not None:
                for floor_name in floor_list:
                    floor_count = row[title_list.index(floor_name)].value
                    print('floor_name = ', floor_name, ' /  floor_count = ', floor_count)
                    if floor_count is not None and int(floor_count) > 0:
                        if 'B' in floor_name:
                            final_floor_name = f'{floor_name}F'
                        elif 'F' in floor_name:
                            floor_name1 = floor_name.replace('F','')
                            final_floor_name = f'{floor_name1}F'
                        elif floor_name == 'P1':
                            final_floor_name = 'RF'
                        elif floor_name == 'P2':
                            final_floor_name = 'PHRF'
                        windows_dict[window_name].append((final_floor_name, floor_count))

            if row[0].value == '' or len(row[0].value) < 2:
                break

            if row[name_index].value is not None:
                # windows_name = f"{row[name_index].value}({row[note_index].value})"
                # windows_dict[row[name_index].value] = row[quantity_index].value
                windows_standard1 = f'{(row[width_index].value):0.3f}'
                windows_standard2 = f'{(row[height_index].value):0.3f}'
                windows_standard3 = f'{(row[area_index].value):0.3f}'
                windows_standard = f"{windows_standard1}*{windows_standard2}={windows_standard3}"
                if row[note_index].value is not None and len(row[note_index].value) > 2:
                    windows_name = f"{row[name_index].value}({row[note_index].value})"
                else:
                    windows_name = f"{row[name_index].value}"

            for floor in windows_dict[window_name]:
                item = ItemStandard(
                    floor=floor[0],
                    location=window_name,
                    roomname=window_name,
                    name=window_name,
                    standard=windows_standard,
                    unit='EA',
                    type='창호',
                    formula=floor[1],
                    sum=floor[1],
                )
                items.append(item)

    # windows
    if '창호산출서' in excel.sheetnames:
        worksheet = excel['창호산출서']

        row_index = 1
        for row in worksheet.iter_rows():
            for column_index, column in enumerate(row):
                if column.value in start_titles:
                    windowslist_row_index = row_index
                    break
            row_index += 1

        for index, column in enumerate(worksheet[windowslist_row_index]):
            match column.value:
                case '부위':
                    part_index = index
                case '품명':
                    name_index = index
                case '규격':
                    standard_index = index
                case '단위':
                    unit_index = index
                case '산식':
                    formular_index = index
                case '물량':
                    quantity_index = index

        for row in worksheet.iter_rows(min_col=0, max_col=worksheet._current_row, min_row=windowslist_row_index+1):
            if (row[part_index].value is not None
                    and row[name_index].value is None
                    and row[standard_index].value is None
                    and row[unit_index].value is None
                    and row[formular_index].value is None
                    and row[quantity_index].value is None
            ):
                try:
                    temp_window_name = row[part_index].value.split('(')[0]
                    if '창호' in temp_window_name:
                        window_name = temp_window_name.split(':')[-1].strip(' ')
                        continue
                except Exception as e:
                    print('예외가 발생했습니다', e)
                    print('창호산출서 : split오류' + str(item))

            # 품명없음 삭제
            if (row[quantity_index].value is None
                    or row[quantity_index].value == "'"
                    or row[quantity_index].value == '0'
                    or row[quantity_index].value == 0
                    or row[quantity_index].value == '물량'
            ):
                continue

            if row[name_index] is not None:
                if window_name in row[name_index].value:
                    continue

            for floor in windows_dict[window_name]:
                window_name_with_floor = '_'.join([window_name, floor[0]])
                item = ItemStandard(
                    floor=floor[0],
                    location=window_name,
                    roomname=window_name,
                    name=row[name_index].value.replace(window_name, window_name_with_floor),
                    standard=row[standard_index].value,
                    unit=row[unit_index].value,
                    type='창호',
                    formula=f"({row[formular_index].value})*<수량>({floor[1]})",
                    sum=float(row[quantity_index].value) * (float(floor[1])),
                )
                items.append(item)




    for item in items:
        delete_duplicate.launch(item)

    items2 = []
    if excel.sheetnames.__contains__('동별집계표'):
        worksheet2 = excel['동별집계표']
        temp_constructionwork = ""
        for row in worksheet2.iter_rows(min_col=1, max_col=5, min_row=4):
            # 중공종
            if (row[1].value.__contains__('내  역  삭  제')
            ):
                continue

            if (row[1].value is not None
                    and row[4].value is None
            ):
                temp_constructionwork = row[1].value.replace(" ", "")

            item2 = ItemStandard2(
                constructionWork=temp_constructionwork,
                name=row[1].value,
                standard=row[2].value,
                unit=row[3].value,
                sum=row[4].value,
            )
            items2.append(item2)

    # for item in items:
    #     print(item)

    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '건축(데이터변경X)'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 30
    new_sheet.column_dimensions["L"].width = 30
    for item in items:
        new_sheet.append(item.to_excel())

    sheet = new_workbook.create_sheet(title='집계표')
    sheet.append(['중공종', '품명', '규격', '단위', '수량(할증전)'])
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30
    for item2 in items2:
        sheet.append(item2.to_excel())

    new_workbook.save(saveFilePath)

    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_normalize('PyCharm')
