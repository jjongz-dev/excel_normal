from openpyxl import load_workbook, Workbook
from ExcelStandard import ExcelStandard
from datetime import datetime
import platform
import subprocess

import ReplaceCivilEarthwork
import ReplaceCivilSidePostPile
import ReplaceCivilCIP
import ReplaceCivilStrut
import ReplaceCivilSGR
import ReplaceCivilLW
import ReplaceCivilRoadDeckingPanel


import re
import ReplacePersonal

fileCreateDate = datetime.strftime(datetime.today(), '%y%m%d_%H%M')
systemOs = platform.system()



fileCreateDate = datetime.strftime(datetime.today(), '%Y%m%d_%H%M')

# 이곳에 현장 폴더명만 변경하면 완료 #######
siteTicketNo = '23-0361_ko'
##################################



if systemOs == 'Darwin':
    openFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/토목.xlsx'
    saveFilePath = '/Users/blue/hb/quantity/' + siteTicketNo + '/end-3-land-' + fileCreateDate + '.xlsx'
else:
    openFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\토목.xlsx'
    saveFilePath = 'D:\\howbuild\\quantity\\'+siteTicketNo+'\토목완성-' + fileCreateDate + '.xlsx'

def excel_normalize(name, column_dimensions=None):
    excel = load_workbook(openFilePath, data_only=True)

    내역목록 = []

    품명확정 = ''
    규격확정 = ''
    부위확정 = ''
    산식확정 = ''
    수량확정 = ''

    if "토목산출서" in excel.sheetnames:
        worksheet = excel['토목산출서']

        for 줄번호, row in enumerate(worksheet.iter_rows(min_row=5)):

            도형값 = row[0].value
            품명값 = row[3].value
            규격값 = row[4].value
            단위값 = row[5].value
            산식값 = row[6].value
            수량값 = row[8].value

            if 단위값 is None or 수량값 == 0 or 수량값 is None:
                continue

            if 품명값 is not None and 단위값 is not None:
                품명확정 = 품명값.replace('\n', '')

            규격확정 = 규격값
            단위확정 = 단위값
            산식확정 = 산식값
            수량확정 = 수량값

            내역 = ExcelStandard(
                층='',
                호='',
                실='',
                대공종='토목',
                중공종='',
                코드='',
                품명=품명확정,
                규격=규격확정,
                단위=단위확정,
                부위='',
                타입='외부',
                산식=산식확정,
                수량=수량확정,
                Remark='',
                개소=''
            )
            내역목록.append(내역)


        # 품명 규격 자동 변경 S #######################
        for 내역 in 내역목록:
            ReplaceCivilEarthwork.launch(내역)
        # 품명 규격 자동 변경 E #######################


    # 저장할 엑셀
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '토목완성'
    head_title = ['층', '호', '실', '대공종', '중공종', '코드', '품명', '규격', '단위', '부위', '타입', '산식', '수량', 'Remark']
    new_sheet.append(head_title)
    new_sheet.column_dimensions["G"].width = 30
    new_sheet.column_dimensions["H"].width = 40

    for 내역 in 내역목록:
        new_sheet.append(내역.to_excel())

    new_workbook.save(saveFilePath)

    # 파싱한 엑셀을 자동으로 띄워서 확인
    systemOs = platform.system()
    if systemOs =='Darwin':
        subprocess.call(['open', saveFilePath])


if __name__ == '__main__':
    excel_normalize('PyCharm')
